# ════════════════════════════════════════════════════════════════════════════
#  ROTATION LOGIC
# ════════════════════════════════════════════════════════════════════════════
import os
import re
import json
import uuid
import datetime
import io

from flask import Blueprint, request, jsonify, send_file
from routes_auth import require_auth, require_admin_role

# ── Blueprint ─────────────────────────────────────────────────────────────
rota_bp = Blueprint('rota', __name__)

# ── Config ────────────────────────────────────────────────────────────────
_BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
ROTA_DIR   = os.path.join(_BASE_DIR, 'rota')
LEAVE_FILE = os.path.join(ROTA_DIR, 'leave_requests.json')
USERS_FILE = os.path.join(_BASE_DIR, 'users.json')

# ── Data helpers ──────────────────────────────────────────────────────────
def _load_json(path: str):
    if not os.path.exists(path):
        return {}
    try:
        with open(path, 'r') as f:
            return json.load(f)
    except Exception as e:
        print(f"[JSON ERROR] {path}: {e}")
        return {}

def _save_json(path: str, data) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w') as f:
        json.dump(data, f, indent=2)

# ── Role helpers ──────────────────────────────────────────────────────────
STAFF_ROLES = {'engineer', 'specialist'}
VALID_LEAVE_TYPES = {'Annual Leave', 'Parental Leave', 'Marital Leave'}

VALID_TRANSITIONS = {
    'Pending':            {'Confirmed', 'Rejected', 'Cancelled'},
    'Confirmed':          {'Withdrawal Pending'},
    'Withdrawal Pending': {'Withdrawn', 'Withdrawal Rejected', 'Cancelled'},
}

def _display_name_from_email(email: str) -> str:
    local = email.split('@')[0]
    parts = re.split(r'[.\-_]', local)
    if not parts:
        return email
    first = parts[0].capitalize()
    if len(parts) >= 2:
        return f"{first} {parts[1][0].upper()}"
    return first

def _get_rota_role(session: dict) -> str:
    role = session.get('role', '')
    if role == 'admin':
        return 'management'
    if role in STAFF_ROLES:
        return 'staff'
    return 'guest'

def _now_iso() -> str:
    return datetime.datetime.utcnow().isoformat(timespec='seconds') + 'Z'

# ── Rotation constants ────────────────────────────────────────────────────
from datetime import date, timedelta

ANCHOR_MONDAY = date(2026, 3, 2)

SPECIALIST_ROTATION = [
    "OFF","OFF","0700-1800","0700-1800","OFF","OFF","OFF",
    "0700-1800","0700-1800","OFF","OFF","0700-1800","0700-1800","0700-1800",
    "OFF","OFF","0900-2000","0900-2000","0900-2000","0900-2000","0900-2000",
    "OFF","OFF","1300-0000","1300-0000","OFF","OFF","OFF",
    "1300-0000","1300-0000","OFF","OFF","1300-0000","1300-0000","1300-0000",
    "1500-0200","1500-0200","OFF","OFF","OFF","1500-0200","1500-0200",
    "OFF","OFF","1500-0200","1500-0200","1500-0200","OFF","OFF",
    "OFF","OFF","2100-0700","2100-0700","2100-0700","2100-0700","2100-0700",
    "OFF","OFF","2100-0700","2100-0700","OFF","OFF","OFF",
    "OFF","OFF","0700-1800","0700-1800","OFF","OFF","OFF",
    "0700-1800","0700-1800","OFF","OFF","0700-1800","0700-1800","0700-1800",
    "OFF","OFF","0900-2000","0900-2000","0900-2000","0900-2000","0900-2000",
    "OFF","OFF","1300-0000","1300-0000","OFF","OFF","OFF",
    "1300-0000","1300-0000","OFF","OFF","1300-0000","1300-0000","1300-0000",
    "1500-0200","1500-0200","OFF","OFF","OFF","1500-0200","1500-0200",
    "OFF","OFF","1500-0200","1500-0200","1500-0200","OFF","OFF",
    "2100-0700","2100-0700","OFF","OFF","2100-0700","2100-0700","2100-0700",
    "2100-0700","2100-0700","OFF","OFF","OFF","OFF","OFF",
]

ENGINEERING_ROTATION = [
    "OFF","0900-1800","0900-1800","0900-1800","0900-1800","OFF","OFF",
    "0900-1800","1000-2000","OFF","OFF","1000-2000","1000-2000","1000-2000",
    "1000-2000","OFF","1000-2000","1000-2000","OFF","OFF","OFF",
]

# ── Person directory ──────────────────────────────────────────────────────
# All identity data (names, rotation membership, HR team) lives in a single
# gitignored JSON keyed by employee_id — the one mandatory, stable,
# collision-free identifier every user has from day one. See
# rota/person_directory.json. Nothing here is committed to source control.
PERSON_DIRECTORY_FILE = os.path.join(ROTA_DIR, 'person_directory.json')
DIRECTORY_AUDIT_FILE  = os.path.join(ROTA_DIR, 'directory_audit_log.json')

VALID_ROTATION_GROUPS = {'management', 'engineering', 'specialist'}

def _load_person_directory() -> dict:
    d = _load_json(PERSON_DIRECTORY_FILE)
    if not isinstance(d, dict) or not d:
        raise RuntimeError(
            f'person_directory.json missing or empty at {PERSON_DIRECTORY_FILE} '
            f'— refusing to start with an empty roster.'
        )
    return d

# rota_label = short grid alias — used as the internal key throughout the
# rota engine, exactly as the old hardcoded names were. Immutable once a
# directory entry is created (see _validate_directory_entry) because every
# transactional file (leave_requests.json, draft_overrides.json,
# cell_notes.json, published_overrides.json) still keys on rota_label, not
# employee_id — renaming it would silently orphan historical records.
# full_name  = legal name for HR exports (PicaPonto, Night Hours) and topbar.
#
# 'active: false' entries are excluded from MANAGEMENT_SHIFTS/
# ENGINEERING_OFFSETS/SPECIALIST_OFFSETS (so they drop out of the rota grid
# and hour computation) but stay in the RID/label/name lookups so historical
# records under their rota_label still resolve to a name instead of a blank.
def _rebuild_person_directory_caches() -> None:
    global _DIR, MANAGEMENT_SHIFTS, ENGINEERING_OFFSETS, SPECIALIST_OFFSETS
    global _RID_TO_LABEL, _LABEL_TO_RID, _LABEL_TO_FULLNAME, _LABEL_TO_HRTEAM

    _DIR = _load_person_directory()
    active = {rid: v for rid, v in _DIR.items() if v.get('active', True)}

    MANAGEMENT_SHIFTS   = {v['rota_label']: v['shift']
                            for v in active.values() if v['rotation_group'] == 'management'}
    ENGINEERING_OFFSETS = {v['rota_label']: v['offset']
                            for v in active.values() if v['rotation_group'] == 'engineering'}
    SPECIALIST_OFFSETS  = {v['rota_label']: v['offset']
                            for v in active.values() if v['rotation_group'] == 'specialist'}

    _RID_TO_LABEL      = {rid: v['rota_label']  for rid, v in _DIR.items()}
    _LABEL_TO_RID      = {v['rota_label']: rid  for rid, v in _DIR.items()}
    _LABEL_TO_FULLNAME = {v['rota_label']: v['full_name'] for v in _DIR.values()}
    _LABEL_TO_HRTEAM   = {v['rota_label']: v.get('hr_team') for v in _DIR.values()}

_rebuild_person_directory_caches()

def _email_to_employee_id() -> dict:
    """Live lookup from the users API payload — never cached/stored, since
    users.json is not ours to manage."""
    users = _load_json(USERS_FILE)
    if not isinstance(users, dict):
        return {}
    return {email: str(info.get('employee_id'))
            for email, info in users.items() if info.get('employee_id')}

def _rota_display_name(email: str) -> str:
    """Short grid label, e.g. 'Joao L'. Falls back to email local-part for
    any account not yet present in person_directory.json."""
    rid = _email_to_employee_id().get(email)
    return _RID_TO_LABEL.get(rid, email.split('@')[0])

def _rota_full_name(email: str) -> str:
    """Full legal name — HR exports and topbar. Same fallback as above."""
    rid = _email_to_employee_id().get(email)
    return _LABEL_TO_FULLNAME.get(_RID_TO_LABEL.get(rid), email.split('@')[0])

def _email_for_rota_name(name: str):
    rid = _LABEL_TO_RID.get(name)
    if not rid:
        return None
    users = _load_json(USERS_FILE)
    if not isinstance(users, dict):
        return None
    for email, info in users.items():
        if str(info.get('employee_id')) == rid:
            return email
    return None

def _hr_teams() -> dict:
    out = {'SOE': [], 'SOS': []}
    for label, team in _LABEL_TO_HRTEAM.items():
        if team in out:
            out[team].append(label)
    return out

def _display_names() -> dict:
    return dict(_LABEL_TO_FULLNAME)

PUBLIC_HOLIDAYS = {
    date(2026,1,1),  date(2026,2,17), date(2026,4,3),
    date(2026,4,5),  date(2026,4,25), date(2026,5,1),
    date(2026,5,12), date(2026,6,4),  date(2026,6,10),
    date(2026,8,15), date(2026,10,5), date(2026,11,1),
    date(2026,12,1), date(2026,12,8), date(2026,12,25),
}

PARENTAL_LEAVE_TYPES    = {"Parental Leave"}
MARITAL_LEAVE_TYPES     = {"Marital Leave"}
AL_APPROVED_STATUSES    = {'Confirmed', 'Withdrawal Pending', 'Withdrawal Rejected'}
AL_PENDING_STATUSES     = {'Pending'}
AL_CLEAR_STATUSES       = {'Rejected', 'Withdrawn', 'Cancelled'}
COVERAGE_REQUIRED_SHIFTS = {'0700-1800', '1500-0200', '2100-0700'}
COVERAGE_FREE_SHIFTS     = {'0900-2000', '1300-0000'}

# ── Shift resolution ──────────────────────────────────────────────────────
def _base_shift(name: str, d: date) -> str:
    delta = (d - ANCHOR_MONDAY).days
    if name in SPECIALIST_OFFSETS:
        idx = (SPECIALIST_OFFSETS[name] + delta) % len(SPECIALIST_ROTATION)
        return SPECIALIST_ROTATION[idx]
    if name in ENGINEERING_OFFSETS:
        idx = (ENGINEERING_OFFSETS[name] + delta) % len(ENGINEERING_ROTATION)
        return ENGINEERING_ROTATION[idx]
    if d.weekday() >= 5 or d in PUBLIC_HOLIDAYS:
        return "OFF"
    return MANAGEMENT_SHIFTS.get(name, "OFF")

def _resolve_shift(name: str, d: date, leave_map: dict,
                   override_map: dict = None) -> str:
    if override_map is not None:
        ov = override_map.get((name, d))
        if ov is not None:
            return ov['shift']

    leave = leave_map.get((name, d))
    if not leave:
        return _base_shift(name, d)

    lt     = leave["leave_type"]
    status = leave["status"]
    base   = _base_shift(name, d)

    if status in AL_CLEAR_STATUSES:
        return base
    if lt in PARENTAL_LEAVE_TYPES:
        return "PARENTAL"
    if lt in MARITAL_LEAVE_TYPES:
        return "MARITAL"
    if status in AL_APPROVED_STATUSES:
        return "AL_APPROVED" if base == "OFF" else f"AL_APPROVED|{base}"
    if status in AL_PENDING_STATUSES:
        return "AL_PENDING" if base == "OFF" else f"AL_PENDING|{base}"
    return base

def _flanking_off_range(person: str, ds: date, de: date) -> tuple:
    """Extend [ds, de] backwards/forwards over consecutive rota-OFF days
    for this person, so a confirmed AL block visually swallows the
    weekends/off-days it's adjacent to. Capped at 14 days each direction."""
    d = ds - timedelta(days=1)
    while _base_shift(person, d) == 'OFF':
        ds = d
        d -= timedelta(days=1)
        if (ds - d).days > 14:
            break
    d = de + timedelta(days=1)
    while _base_shift(person, d) == 'OFF':
        de = d
        d += timedelta(days=1)
        if (d - de).days > 14:
            break
    return ds, de

def _build_leave_map(leave_list: list) -> dict:
    lmap = {}
    for r in leave_list:
        try:
            ds = date.fromisoformat(r["date_start"])
            de = date.fromisoformat(r["date_end"])
        except (KeyError, ValueError):
            continue
        # Only expand over flanking OFF days once AL is actually confirmed
        # (or in a state that was previously confirmed). Pending/provisional
        # requests show exactly the days requested, nothing more.
        if r.get("leave_type") == "Annual Leave" and r.get("status") in AL_APPROVED_STATUSES:
            ds, de = _flanking_off_range(r["name"], ds, de)
        d = ds
        while d <= de:
            lmap[(r["name"], d)] = {
                "leave_type": r["leave_type"],
                "status":     r["status"],
            }
            d += timedelta(days=1)
    return lmap

def _build_override_map(overrides: list) -> dict:
    omap = {}
    for o in overrides:
        try:
            d = date.fromisoformat(o['date'])
        except (KeyError, ValueError):
            continue
        omap[(o['person'], d)] = o
    return omap

def _build_note_map(notes: list) -> dict:
    """Keys: (person, date) → note text."""
    nmap = {}
    for n in notes:
        try:
            d = date.fromisoformat(n['date'])
        except (KeyError, ValueError):
            continue
        nmap[(n['person'], d)] = n.get('note', '')
    return nmap

# ── Schedule builder (shared by published + draft routes) ─────────────────
def _build_schedule(date_from: date, date_to: date,
                    leave_map: dict, override_map: dict,
                    note_map: dict) -> list:
    today = date.today()
    days  = []
    d     = date_from
    while d <= date_to:
        day = {
            'date':              d.isoformat(),
            'weekday':           d.strftime('%A'),
            'is_today':          d == today,
            'is_weekend':        d.weekday() >= 5,
            'is_public_holiday': d in PUBLIC_HOLIDAYS,
            'shifts':            {},
        }
        for name in list(MANAGEMENT_SHIFTS) + list(ENGINEERING_OFFSETS) + list(SPECIALIST_OFFSETS):
            team = ('Management'  if name in MANAGEMENT_SHIFTS  else
                    'Engineering' if name in ENGINEERING_OFFSETS else
                    'Specialists')
            shift = _resolve_shift(name, d, leave_map, override_map)
            note  = note_map.get((name, d))
            day['shifts'][name] = {
                'team':  team,
                'shift': shift,
                'note':  note,
            }
        days.append(day)
        d += timedelta(days=1)
    return days

# ════════════════════════════════════════════════════════════════════════════
#  FILE PATHS
# ════════════════════════════════════════════════════════════════════════════
CONFIG_FILE              = os.path.join(ROTA_DIR, 'config.json')
DRAFT_FILE               = os.path.join(ROTA_DIR, 'draft_overrides.json')
DRAFT_LOCK_FILE          = os.path.join(ROTA_DIR, 'draft_lock.json')
PUBLISHED_OVERRIDES_FILE = os.path.join(ROTA_DIR, 'published_overrides.json')
CELL_NOTES_FILE          = os.path.join(ROTA_DIR, 'cell_notes.json')
FEEDBACK_FILE            = os.path.join(ROTA_DIR, 'feedback.json')
HOURS_POT_FILE           = os.path.join(ROTA_DIR, 'hours_pot.json')
AL_ALLOWANCE_FILE        = os.path.join(ROTA_DIR, 'al_allowance.json')

# ── Config ────────────────────────────────────────────────────────────────
DEFAULT_CONFIG = {
    'next_year_open_from': '11-01',
    'custom_shift_colors': [],
    'custom_shift_color_map': {},
}

def _load_config() -> dict:
    cfg = _load_json(CONFIG_FILE)
    if not isinstance(cfg, dict):
        cfg = {}
    return {**DEFAULT_CONFIG, **cfg}

# ── Notes ──────────────────────────────────────────────────────────────────
def _load_notes() -> list:
    data = _load_json(CELL_NOTES_FILE)
    return data if isinstance(data, list) else []

def _save_notes(notes: list) -> None:
    _save_json(CELL_NOTES_FILE, notes)

# ── Draft helpers ──────────────────────────────────────────────────────────
DRAFT_LOCK_TIMEOUT_MIN = 240

def _load_draft_overrides() -> list:
    data = _load_json(DRAFT_FILE)
    return data if isinstance(data, list) else []

def _save_draft_overrides(overrides: list) -> None:
    _save_json(DRAFT_FILE, overrides)

def _load_draft_lock():
    data = _load_json(DRAFT_LOCK_FILE)
    if not isinstance(data, dict) or not data.get('locked_by'):
        return None
    try:
        locked_at = datetime.datetime.fromisoformat(
            data['locked_at'].replace('Z', '+00:00'))
        age_min = (datetime.datetime.now(datetime.timezone.utc)
                   - locked_at).total_seconds() / 60
        if age_min > DRAFT_LOCK_TIMEOUT_MIN:
            return None
    except Exception:
        pass
    return data

def _save_draft_lock(username: str, name: str) -> dict:
    lock = {'locked_by': username, 'locked_by_name': name,
            'locked_at': _now_iso()}
    _save_json(DRAFT_LOCK_FILE, lock)
    return lock

def _clear_draft_lock() -> None:
    _save_json(DRAFT_LOCK_FILE, {})

def _require_management():
    if _get_rota_role(request.session) != 'management':
        return jsonify({'ok': False, 'error': 'Not authorised'}), 403
    return None

def _require_draft_lock_held_by_me():
    session = request.session
    lock    = _load_draft_lock()
    if not lock or lock.get('locked_by') != session['username']:
        return jsonify({
            'ok': False,
            'error': 'You do not currently hold the draft lock.',
        }), 409
    return None

# ── AL bundling helpers ────────────────────────────────────────────────────
def _bundle_al_overrides(al_overrides: list, person: str) -> list:
    """
    Given a list of al_toggle override records for one person,
    group consecutive dates into bundles and extend each bundle
    to cover flanking OFF days (based on base rotation).
    Returns list of (date_start, date_end, shift_code) tuples.
    'shift_code' is AL_APPROVED or AL_PENDING from the override.
    """
    if not al_overrides:
        return []

    # Sort by date
    sorted_ovs = sorted(al_overrides, key=lambda o: o['date'])
    dates = [date.fromisoformat(o['date']) for o in sorted_ovs]

    # Get the AL type from the first override (all in a bundle share type)
    def _al_type(shift: str) -> str:
        if shift.startswith('AL_APPROVED'):
            return 'AL_APPROVED'
        return 'AL_PENDING'

    # Group into consecutive runs
    groups = []
    current = [sorted_ovs[0]]
    for i in range(1, len(sorted_ovs)):
        prev_d = date.fromisoformat(sorted_ovs[i-1]['date'])
        curr_d = date.fromisoformat(sorted_ovs[i]['date'])
        if (curr_d - prev_d).days == 1:
            current.append(sorted_ovs[i])
        else:
            groups.append(current)
            current = [sorted_ovs[i]]
    groups.append(current)

    bundles = []
    for group in groups:
        ds = date.fromisoformat(group[0]['date'])
        de = date.fromisoformat(group[-1]['date'])
        al_code = _al_type(group[0]['shift'])

        # Extend backwards over flanking OFF days
        d = ds - timedelta(days=1)
        while _base_shift(person, d) == 'OFF':
            ds = d
            d -= timedelta(days=1)
            if (ds - d).days > 14:  # safety cap
                break

        # Extend forwards over flanking OFF days
        d = de + timedelta(days=1)
        while _base_shift(person, d) == 'OFF':
            de = d
            d += timedelta(days=1)
            if (d - de).days > 14:
                break

        bundles.append((ds, de, al_code))

    return bundles

# ════════════════════════════════════════════════════════════════════════════
#  ROUTES
# ════════════════════════════════════════════════════════════════════════════

@rota_bp.route('/rota/me', methods=['GET'])
@require_auth
def rota_me():
    session   = request.session
    rota_role = _get_rota_role(session)
    username  = session['username']
    name      = _rota_display_name(username)
    full_name = _rota_full_name(username)
    role      = session.get('role', '')
    team      = ('Engineering' if role == 'engineer' else
                 'Specialists' if role == 'specialist' else
                 'Management'  if role == 'admin' else None)
    return jsonify({'ok': True, 'username': username,
                    'rota_role': rota_role, 'name': name,
                    'full_name': full_name, 'team': team})


@rota_bp.route('/rota/members', methods=['GET'])
@require_auth
def rota_members():
    users = _load_json(USERS_FILE)
    if not isinstance(users, dict):
        return jsonify({'ok': False, 'error': 'Could not load users'}), 500
    email_to_rid = {email: str(info.get('employee_id'))
                     for email, info in users.items() if info.get('employee_id')}
    result = {}
    for email, info in users.items():
        role  = info.get('role', '')
        team  = ('Engineering' if role == 'engineer' else
                 'Specialists' if role == 'specialist' else
                 'Management'  if role == 'admin' else 'Other')
        label = _RID_TO_LABEL.get(email_to_rid.get(email), email.split('@')[0])
        result[email] = {
            'name':      label,
            'full_name': _LABEL_TO_FULLNAME.get(label, label),
            'team':      team,
            'role':      role,
        }
    return jsonify({'ok': True, 'members': result})


@rota_bp.route('/rota/roster', methods=['GET'])
@require_auth
def rota_roster():
    """Roster grouped by rotation team, in person_directory.json order —
    replaces the MGMT_NAMES/ENG_NAMES literals that used to live in the
    frontend."""
    return jsonify({
        'ok': True,
        'management':  list(MANAGEMENT_SHIFTS.keys()),
        'engineering': list(ENGINEERING_OFFSETS.keys()),
        'specialists': list(SPECIALIST_OFFSETS.keys()),
    })


# ── Directory admin (add / edit / hide / remove people) ────────────────────
# Management only. Every write is appended to a small audit log — this data
# drives shift computation and payroll exports, so unlike leave requests
# (which already carry full history) a bad edit here has a much larger
# blast radius and deserves a paper trail even without full undo.

def _load_directory_audit() -> list:
    d = _load_json(DIRECTORY_AUDIT_FILE)
    return d if isinstance(d, list) else []

def _append_directory_audit(action: str, employee_id: str, by: str,
                            before, after) -> None:
    log = _load_directory_audit()
    log.append({
        'at':          _now_iso(),
        'by':          by,
        'action':      action,        # 'create' | 'update' | 'delete'
        'employee_id': employee_id,
        'before':      before,
        'after':       after,
    })
    _save_json(DIRECTORY_AUDIT_FILE, log)

def _validate_directory_entry(data: dict, existing_rid=None):
    """Returns (cleaned_entry, error_message_or_None). existing_rid is set
    when editing — used only to exclude the entry itself from the
    rota_label collision check, never to allow changing its own label."""
    full_name      = (data.get('full_name') or '').strip()
    rota_label     = (data.get('rota_label') or '').strip()
    rotation_group = (data.get('rotation_group') or '').strip()
    hr_team        = data.get('hr_team') or None
    soe_join_date  = data.get('soe_join_date') or None
    active         = bool(data.get('active', True))

    if not full_name:
        return None, 'full_name is required'
    if not rota_label:
        return None, 'rota_label is required'
    if rotation_group not in VALID_ROTATION_GROUPS:
        return None, f'rotation_group must be one of {sorted(VALID_ROTATION_GROUPS)}'
    if hr_team not in (None, 'SOE', 'SOS'):
        return None, 'hr_team must be SOE, SOS, or null'
    if soe_join_date:
        try:
            date.fromisoformat(soe_join_date)
        except ValueError:
            return None, 'soe_join_date must be YYYY-MM-DD'

    for rid, v in _DIR.items():
        if rid == existing_rid:
            continue
        if v['rota_label'].strip().lower() == rota_label.lower():
            return None, (f"rota_label '{rota_label}' is already used by "
                          f"employee_id {rid} ({v['full_name']})")

    entry = {
        'full_name':      full_name,
        'rota_label':     rota_label,
        'rotation_group': rotation_group,
        'hr_team':        hr_team,
        'soe_join_date':  soe_join_date,
        'active':         active,
    }
    if rotation_group == 'management':
        shift = (data.get('shift') or '').strip()
        if not re.match(r'^\d{4}-\d{4}$', shift):
            return None, 'shift must be HHMM-HHMM (e.g. 0900-1730) for management'
        entry['shift'] = shift
    else:
        try:
            offset = int(data.get('offset'))
        except (TypeError, ValueError):
            return None, 'offset must be an integer for engineering/specialist'
        if offset < 0:
            return None, 'offset must be 0 or greater'
        entry['offset'] = offset

    return entry, None


@rota_bp.route('/rota/directory', methods=['GET'])
@require_auth
def rota_directory_get():
    """Full directory, active and hidden alike, for the admin table.
    Each entry is annotated with the resolved login email when one exists,
    so the admin can see who's actually linked to a login account."""
    err = _require_management()
    if err: return err
    users = _load_json(USERS_FILE)
    if not isinstance(users, dict):
        users = {}
    rid_to_email = {}
    for email, info in users.items():
        rid = str(info.get('employee_id') or '')
        if rid:
            rid_to_email[rid] = email
    out = {rid: {**v, 'email': rid_to_email.get(rid)} for rid, v in _DIR.items()}
    return jsonify({'ok': True, 'directory': out})


@rota_bp.route('/rota/directory', methods=['POST'])
@require_auth
def rota_directory_post():
    err = _require_management()
    if err: return err
    session     = request.session
    data        = request.get_json(silent=True) or {}
    employee_id = str(data.get('employee_id') or '').strip()

    if not employee_id.isdigit():
        return jsonify({'ok': False, 'error': 'employee_id must be a positive integer'}), 400
    if employee_id in _DIR:
        return jsonify({'ok': False, 'error': f'employee_id {employee_id} already exists'}), 409

    entry, err_msg = _validate_directory_entry(data)
    if err_msg:
        return jsonify({'ok': False, 'error': err_msg}), 400

    directory = _load_json(PERSON_DIRECTORY_FILE)
    if not isinstance(directory, dict):
        directory = {}
    directory[employee_id] = entry
    _save_json(PERSON_DIRECTORY_FILE, directory)
    _rebuild_person_directory_caches()
    _append_directory_audit('create', employee_id, session['username'], None, entry)

    return jsonify({'ok': True, 'employee_id': employee_id, 'entry': entry})


@rota_bp.route('/rota/directory/<employee_id>', methods=['PUT'])
@require_auth
def rota_directory_put(employee_id):
    """Edit an existing entry, or flip 'active' to hide/show them on the
    rota. rota_label and employee_id cannot be changed here — see the
    module comment above _rebuild_person_directory_caches for why."""
    err = _require_management()
    if err: return err
    if employee_id not in _DIR:
        return jsonify({'ok': False, 'error': 'employee_id not found'}), 404

    session  = request.session
    data     = dict(request.get_json(silent=True) or {})
    before   = dict(_DIR[employee_id])

    # rota_label is immutable post-creation — ignore any attempt to change
    # it rather than erroring, so a naive "resubmit the whole form" edit
    # from the UI doesn't fail just because the label field round-tripped.
    data['rota_label'] = before['rota_label']

    entry, err_msg = _validate_directory_entry(data, existing_rid=employee_id)
    if err_msg:
        return jsonify({'ok': False, 'error': err_msg}), 400

    directory = _load_json(PERSON_DIRECTORY_FILE)
    if not isinstance(directory, dict):
        directory = {}
    directory[employee_id] = entry
    _save_json(PERSON_DIRECTORY_FILE, directory)
    _rebuild_person_directory_caches()
    _append_directory_audit('update', employee_id, session['username'], before, entry)

    return jsonify({'ok': True, 'employee_id': employee_id, 'entry': entry})


@rota_bp.route('/rota/directory/<employee_id>', methods=['DELETE'])
@require_auth
def rota_directory_delete(employee_id):
    """Hard delete. Prefer hiding (PUT with active:false) — deleting removes
    this rota_label from every name-resolution lookup, so historical leave
    requests, overrides, and notes under that label will no longer show a
    resolved name anywhere in the UI. The frontend warns about this before
    calling here; the backend does not re-check usage across those files."""
    err = _require_management()
    if err: return err
    directory = _load_json(PERSON_DIRECTORY_FILE)
    if not isinstance(directory, dict) or employee_id not in directory:
        return jsonify({'ok': False, 'error': 'employee_id not found'}), 404

    session = request.session
    before  = directory.pop(employee_id)
    _save_json(PERSON_DIRECTORY_FILE, directory)
    _rebuild_person_directory_caches()
    _append_directory_audit('delete', employee_id, session['username'], before, None)

    return jsonify({'ok': True})


@rota_bp.route('/rota/directory/audit', methods=['GET'])
@require_auth
def rota_directory_audit_get():
    err = _require_management()
    if err: return err
    log = _load_directory_audit()
    return jsonify({'ok': True, 'log': list(reversed(log))[:200]})


@rota_bp.route('/rota/schedule', methods=['GET'])
@require_auth
def rota_schedule():
    rota_role = _get_rota_role(request.session)
    try:
        date_from = date.fromisoformat(
            request.args.get('from', date.today().isoformat()))
        date_to   = date.fromisoformat(
            request.args.get('to', (date.today() + timedelta(weeks=5)).isoformat()))
    except ValueError:
        return jsonify({'ok': False,
                        'error': 'Invalid date format, use YYYY-MM-DD'}), 400

    if rota_role != 'management':
        max_to = date.today() + timedelta(weeks=5)
        if date_to > max_to:
            date_to = max_to

    leave_list = _load_json(LEAVE_FILE)
    if not isinstance(leave_list, list):
        leave_list = []

    published_overrides = _load_json(PUBLISHED_OVERRIDES_FILE)
    if not isinstance(published_overrides, list):
        published_overrides = []

    notes = _load_notes()

    leave_map    = _build_leave_map(leave_list)
    override_map = _build_override_map(published_overrides)
    note_map     = _build_note_map(notes)

    days = _build_schedule(date_from, date_to, leave_map, override_map, note_map)
    return jsonify({'ok': True, 'days': days})


def _print_month_allowed_for_staff(year: int, month: int) -> bool:
    """Staff/non-management may only export the current calendar month,
    or next month once within the final 10 days of the current month."""
    today = date.today()
    if (year, month) == (today.year, today.month):
        return True
    if today.month == 12:
        next_year, next_month = today.year + 1, 1
        last_day_current = date(today.year, 12, 31)
    else:
        next_year, next_month = today.year, today.month + 1
        last_day_current = date(next_year, next_month, 1) - timedelta(days=1)
    if (year, month) == (next_year, next_month):
        return (last_day_current - today).days <= 10
    return False


@rota_bp.route('/rota/print-export', methods=['GET'])
@require_auth
def rota_print_export():
    """Full-month schedule for the print/PDF export. Management: any month.
    Staff: current month, or next month within its final 10 days. Guests: none."""
    rota_role = _get_rota_role(request.session)
    if rota_role == 'guest':
        return jsonify({'ok': False, 'error': 'Not authorised'}), 403

    month_param = request.args.get('month', '')  # YYYY-MM
    try:
        year, month = [int(x) for x in month_param.split('-')]
        date_from = date(year, month, 1)
        date_to   = (date(year, month + 1, 1) - timedelta(days=1)) if month < 12 \
                    else date(year, 12, 31)
    except (ValueError, AttributeError):
        return jsonify({'ok': False, 'error': 'month must be YYYY-MM'}), 400

    if rota_role != 'management' and not _print_month_allowed_for_staff(year, month):
        return jsonify({'ok': False,
                        'error': 'This month is not yet available for export.'}), 403

    leave_list = _load_json(LEAVE_FILE)
    if not isinstance(leave_list, list):
        leave_list = []
    published_overrides = _load_json(PUBLISHED_OVERRIDES_FILE)
    if not isinstance(published_overrides, list):
        published_overrides = []
    notes = _load_notes()

    leave_map    = _build_leave_map(leave_list)
    override_map = _build_override_map(published_overrides)
    note_map     = _build_note_map(notes)

    days = _build_schedule(date_from, date_to, leave_map, override_map, note_map)
    return jsonify({'ok': True, 'days': days, 'month': month_param})


@rota_bp.route('/rota/leave', methods=['GET'])
@require_auth
def rota_leave_get():
    session    = request.session
    rota_role  = _get_rota_role(session)
    username   = session['username']
    leave_list = _load_json(LEAVE_FILE)
    if not isinstance(leave_list, list):
        leave_list = []

    if rota_role == 'management':
        return jsonify({'ok': True, 'leave': leave_list})

    my_leave = [
        r for r in leave_list
        if r.get('username') == username or
           (r.get('on_behalf') and r.get('username') == username)
    ]
    return jsonify({'ok': True, 'leave': my_leave})


@rota_bp.route('/rota/leave', methods=['POST'])
@require_auth
def rota_leave_post():
    session   = request.session
    rota_role = _get_rota_role(session)

    if rota_role == 'guest':
        return jsonify({'ok': False, 'error': 'Not authorised'}), 403

    data       = request.get_json(silent=True) or {}
    date_start = data.get('date_start', '').strip()
    date_end   = data.get('date_end', '').strip()
    leave_type = data.get('leave_type', '').strip()
    on_behalf  = data.get('on_behalf', False)
    target     = data.get('target_username', '').strip()

    if not date_start or not date_end or not leave_type:
        return jsonify({'ok': False, 'error': 'Missing fields'}), 400
    if leave_type not in VALID_LEAVE_TYPES:
        return jsonify({'ok': False, 'error': 'Invalid leave type'}), 400
    if date_end < date_start:
        return jsonify({'ok': False, 'error': 'End date before start date'}), 400

    bypass = bool(data.get('bypass_blocker', False))
    if not bypass:
        cfg       = _load_config()
        today     = date.today()
        next_year = today.year + 1
        try:
            mm, dd    = cfg['next_year_open_from'].split('-')
            open_date = date(today.year, int(mm), int(dd))
        except (ValueError, KeyError):
            open_date = date(today.year, 11, 1)
        try:
            ds = date.fromisoformat(date_start)
        except ValueError:
            ds = None
        if ds and ds.year == next_year and today < open_date:
            return jsonify({
                'ok': False,
                'error': f'Leave requests for {next_year} open on '
                         f'{open_date.strftime("%d-%m-%Y")}',
                'blocked':   True,
                'open_date': open_date.isoformat(),
            }), 400

    if on_behalf:
        if rota_role != 'management':
            return jsonify({'ok': False,
                            'error': 'Not authorised for on-behalf requests'}), 403
        if not target:
            return jsonify({'ok': False,
                            'error': 'target_username required for on-behalf'}), 400
        username = target
    else:
        username = session['username']

    name = _rota_display_name(username)
    leave_list = _load_json(LEAVE_FILE)
    if not isinstance(leave_list, list):
        leave_list = []

    leave_list.append({
        'id':          str(uuid.uuid4())[:8],
        'name':        name,
        'username':    username,
        'on_behalf':   on_behalf,
        'created_by':  session['username'],
        'created_at':  _now_iso(),
        'date_start':  date_start,
        'date_end':    date_end,
        'leave_type':  leave_type,
        'status':      'Pending',
        'actioned_by': None,
        'actioned_at': None,
        'history': [{'status': 'Pending', 'by': session['username'],
                     'at': _now_iso()}],
    })
    _save_json(LEAVE_FILE, leave_list)
    return jsonify({'ok': True})


@rota_bp.route('/rota/leave/<leave_id>', methods=['PUT'])
@require_auth
def rota_leave_put(leave_id):
    session    = request.session
    rota_role  = _get_rota_role(session)
    data       = request.get_json(silent=True) or {}
    new_status = data.get('status', '').strip()

    leave_list = _load_json(LEAVE_FILE)
    if not isinstance(leave_list, list):
        return jsonify({'ok': False, 'error': 'No data'}), 500

    idx = next((i for i, r in enumerate(leave_list)
                if r.get('id') == leave_id), None)
    if idx is None:
        return jsonify({'ok': False, 'error': 'Not found'}), 404

    entry          = leave_list[idx]
    current_status = entry.get('status', '')
    mgmt_force     = bool(data.get('mgmt_force', False))
    mgmt_reinstate = bool(data.get('mgmt_reinstate', False))

    # Management-only bypass transitions
    if rota_role == 'management':
        if mgmt_force and current_status == 'Confirmed' and new_status == 'Withdrawal Pending':
            pass  # allowed — skip normal transition check
        elif mgmt_reinstate and current_status == 'Withdrawn' and new_status == 'Confirmed':
            pass  # reinstate a withdrawn entry
        else:
            allowed = VALID_TRANSITIONS.get(current_status, set())
            if new_status not in allowed:
                return jsonify({'ok': False,
                                'error': f'Cannot transition from {current_status} '
                                         f'to {new_status}'}), 400
    else:
        allowed = VALID_TRANSITIONS.get(current_status, set())
        if new_status not in allowed:
            return jsonify({'ok': False,
                            'error': f'Cannot transition from {current_status} '
                                     f'to {new_status}'}), 400
        SELF_SERVICE = {'Withdrawal Pending', 'Cancelled'}
        if new_status not in SELF_SERVICE:
            return jsonify({'ok': False, 'error': 'Not authorised'}), 403
        if entry.get('username') != session['username']:
            return jsonify({'ok': False, 'error': 'Not authorised'}), 403

    now = _now_iso()
    entry['status']      = new_status
    entry['actioned_by'] = session['username']
    entry['actioned_at'] = now
    if 'history' not in entry:
        entry['history'] = []
    entry['history'].append({'status': new_status, 'by': session['username'],
                             'at': now})
    leave_list[idx] = entry
    _save_json(LEAVE_FILE, leave_list)
    return jsonify({'ok': True})


# ── Config routes ─────────────────────────────────────────────────────────

@rota_bp.route('/rota/config', methods=['GET'])
@require_auth
def rota_config_get():
    return jsonify({'ok': True, 'config': _load_config()})


@rota_bp.route('/rota/config', methods=['PUT'])
@require_auth
def rota_config_put():
    if _get_rota_role(request.session) != 'management':
        return jsonify({'ok': False, 'error': 'Not authorised'}), 403
    data = request.get_json(silent=True) or {}
    cfg  = _load_config()
    if 'next_year_open_from' in data:
        val = data['next_year_open_from'].strip()
        try:
            datetime.datetime.strptime(val, '%m-%d')
        except ValueError:
            return jsonify({'ok': False,
                            'error': 'Invalid date format, use MM-DD'}), 400
        cfg['next_year_open_from'] = val
    if 'custom_shift_colors' in data:
        colors = data['custom_shift_colors']
        if isinstance(colors, list):
            cfg['custom_shift_colors'] = colors[-5:]  # keep last 5
    if 'custom_shift_color_map' in data:
        color_map = data['custom_shift_color_map']
        if isinstance(color_map, dict):
            existing = cfg.get('custom_shift_color_map', {})
            if not isinstance(existing, dict):
                existing = {}
            existing.update({
                str(k): str(v) for k, v in color_map.items()
                if isinstance(k, str) and isinstance(v, str)
            })
            cfg['custom_shift_color_map'] = existing
    _save_json(CONFIG_FILE, cfg)
    return jsonify({'ok': True, 'config': cfg})


# ── Cell notes routes ──────────────────────────────────────────────────────

@rota_bp.route('/rota/note', methods=['PUT'])
@require_auth
def rota_note_put():
    """Add or update a note on any cell. Management only. Works outside
    draft mode — notes are independent of shift overrides."""
    err = _require_management()
    if err: return err

    session = request.session
    data    = request.get_json(silent=True) or {}
    person  = data.get('person', '').strip()
    date_s  = data.get('date', '').strip()
    note    = data.get('note', '').strip()

    if not person or not date_s:
        return jsonify({'ok': False, 'error': 'person and date required'}), 400
    try:
        date.fromisoformat(date_s)
    except ValueError:
        return jsonify({'ok': False, 'error': 'Invalid date format'}), 400

    notes = _load_notes()
    existing = next((n for n in notes
                     if n['person'] == person and n['date'] == date_s), None)
    now = _now_iso()
    if note:
        if existing:
            existing.update({'note': note, 'updated_by': session['username'],
                             'updated_at': now})
        else:
            notes.append({'id': str(uuid.uuid4())[:8], 'person': person,
                          'date': date_s, 'note': note,
                          'created_by': session['username'], 'created_at': now})
    else:
        # Empty note = delete
        notes = [n for n in notes
                 if not (n['person'] == person and n['date'] == date_s)]

    _save_notes(notes)
    return jsonify({'ok': True})


@rota_bp.route('/rota/note', methods=['DELETE'])
@require_auth
def rota_note_delete():
    err = _require_management()
    if err: return err
    data   = request.get_json(silent=True) or {}
    person = data.get('person', '').strip()
    date_s = data.get('date', '').strip()
    if not person or not date_s:
        return jsonify({'ok': False, 'error': 'person and date required'}), 400
    notes = _load_notes()
    notes = [n for n in notes
             if not (n['person'] == person and n['date'] == date_s)]
    _save_notes(notes)
    return jsonify({'ok': True})


# ── Draft routes ───────────────────────────────────────────────────────────

@rota_bp.route('/rota/draft/status', methods=['GET'])
@require_auth
def rota_draft_status():
    err = _require_management()
    if err: return err
    return jsonify({'ok': True, 'lock': _load_draft_lock()})


@rota_bp.route('/rota/draft/lock', methods=['POST'])
@require_auth
def rota_draft_lock():
    err = _require_management()
    if err: return err
    session = request.session
    lock    = _load_draft_lock()
    if lock and lock.get('locked_by') != session['username']:
        return jsonify({
            'ok': False,
            'error': f"Draft is currently locked by "
                     f"{lock.get('locked_by_name', lock.get('locked_by'))}",
            'lock': lock,
        }), 409
    name = _rota_display_name(session['username'])
    lock = _save_draft_lock(session['username'], name)
    return jsonify({'ok': True, 'lock': lock})


@rota_bp.route('/rota/draft/unlock', methods=['POST'])
@require_auth
def rota_draft_unlock():
    err = _require_management()
    if err: return err
    session = request.session
    data    = request.get_json(silent=True) or {}
    force   = bool(data.get('force', False))
    lock    = _load_draft_lock()
    if not lock:
        return jsonify({'ok': True})
    if lock.get('locked_by') != session['username'] and not force:
        return jsonify({'ok': False,
                        'error': 'Draft is locked by another user'}), 403
    _clear_draft_lock()
    return jsonify({'ok': True})


@rota_bp.route('/rota/draft', methods=['GET'])
@require_auth
def rota_draft_get():
    err = _require_management()
    if err: return err

    try:
        date_from = date.fromisoformat(
            request.args.get('from', date.today().isoformat()))
        date_to   = date.fromisoformat(
            request.args.get('to', (date.today() + timedelta(weeks=8)).isoformat()))
    except ValueError:
        return jsonify({'ok': False,
                        'error': 'Invalid date format, use YYYY-MM-DD'}), 400

    overrides = _load_draft_overrides()
    leave_list = _load_json(LEAVE_FILE)
    if not isinstance(leave_list, list):
        leave_list = []
    published_overrides = _load_json(PUBLISHED_OVERRIDES_FILE)
    if not isinstance(published_overrides, list):
        published_overrides = []
    notes = _load_notes()

    leave_map    = _build_leave_map(leave_list)
    override_map = _build_override_map(published_overrides)
    override_map.update(_build_override_map(overrides))
    note_map     = _build_note_map(notes)

    days = _build_schedule(date_from, date_to, leave_map, override_map, note_map)
    return jsonify({'ok': True, 'overrides': overrides, 'days': days})


@rota_bp.route('/rota/draft/override', methods=['PUT'])
@require_auth
def rota_draft_override_put():
    err = _require_management()
    if err: return err
    err = _require_draft_lock_held_by_me()
    if err: return err

    session = request.session
    data    = request.get_json(silent=True) or {}
    person  = data.get('person', '').strip()
    date_s  = data.get('date', '').strip()
    shift   = data.get('shift', '').strip() if data.get('shift') else None
    note    = data.get('note', '').strip() if data.get('note') else None
    ov_type = data.get('type', 'shift_change').strip()

    if not person or not date_s or (not shift and ov_type != 'revert_to_original'):
        return jsonify({'ok': False,
                        'error': 'person, date and shift are required'}), 400
    try:
        d = date.fromisoformat(date_s)
    except ValueError:
        return jsonify({'ok': False, 'error': 'Invalid date format'}), 400
    
    if ov_type == 'revert_to_original':
        shift = _base_shift(person, d)

    overrides = _load_draft_overrides()
    leave_list = _load_json(LEAVE_FILE)
    if not isinstance(leave_list, list):
        leave_list = []
    leave_map = _build_leave_map(leave_list)
    published_overrides = _load_json(PUBLISHED_OVERRIDES_FILE)
    if not isinstance(published_overrides, list):
        published_overrides = []
    published_map = _build_override_map(published_overrides)

    existing = next((o for o in overrides
                     if o['person'] == person and o['date'] == date_s), None)
    if existing:
        previous_shift = existing.get('previous_shift')
        existing.update({'shift': shift, 'note': note, 'type': ov_type,
                         'updated_by': session['username'],
                         'updated_at': _now_iso()})
    else:
        previous_shift = _resolve_shift(person, d, leave_map, published_map)
        overrides.append({
            'id':             str(uuid.uuid4())[:8],
            'person':         person,
            'date':           date_s,
            'shift':          shift,
            'previous_shift': previous_shift,
            'note':           note,
            'type':           ov_type,
            'created_by':     session['username'],
            'created_at':     _now_iso(),
        })

    _save_draft_overrides(overrides)
    return jsonify({'ok': True, 'overrides': overrides})


@rota_bp.route('/rota/draft/override/<override_id>', methods=['DELETE'])
@require_auth
def rota_draft_override_delete(override_id):
    err = _require_management()
    if err: return err
    err = _require_draft_lock_held_by_me()
    if err: return err

    overrides     = _load_draft_overrides()
    new_overrides = [o for o in overrides if o.get('id') != override_id]
    if len(new_overrides) == len(overrides):
        return jsonify({'ok': False, 'error': 'Override not found'}), 404
    _save_draft_overrides(new_overrides)
    return jsonify({'ok': True, 'overrides': new_overrides})


@rota_bp.route('/rota/draft/discard', methods=['POST'])
@require_auth
def rota_draft_discard():
    err = _require_management()
    if err: return err
    session = request.session
    lock    = _load_draft_lock()
    if lock and lock.get('locked_by') == session['username']:
        _clear_draft_lock()
    return jsonify({'ok': True})


@rota_bp.route('/rota/draft/publish', methods=['POST'])
@require_auth
def rota_draft_publish():
    err = _require_management()
    if err: return err
    err = _require_draft_lock_held_by_me()
    if err: return err

    session   = request.session
    overrides = _load_draft_overrides()

    if not overrides:
        _clear_draft_lock()
        return jsonify({'ok': True, 'published': 0,
                        'al_created': 0, 'shift_applied': 0,
                        'warnings': []})

    published_overrides = _load_json(PUBLISHED_OVERRIDES_FILE)
    if not isinstance(published_overrides, list):
        published_overrides = []

    leave_list = _load_json(LEAVE_FILE)
    if not isinstance(leave_list, list):
        leave_list = []

    al_created    = 0
    shift_applied = 0
    warnings      = []
    now           = _now_iso()
    today         = date.today()
    five_week_end = today + timedelta(weeks=5)

    # ── Separate al_toggle overrides by person ────────────────────────────
    al_by_person   = {}
    other_overrides = []
    for ov in overrides:
        if ov.get('type') == 'al_toggle':
            p = ov['person']
            al_by_person.setdefault(p, []).append(ov)
        else:
            other_overrides.append(ov)

    # ── Process AL overrides — bundle per person ──────────────────────────
    for person, al_ovs in al_by_person.items():
        # Separate adds (AL_APPROVED / AL_PENDING) from removes (revert to base)
        add_ovs    = [o for o in al_ovs if 'AL_' in o['shift']]
        remove_ovs = [o for o in al_ovs if 'AL_' not in o['shift']]

        email    = _email_for_rota_name(person) or ''

        # ── Handle AL additions — bundle consecutive days ─────────────────
        if add_ovs:
            bundles = _bundle_al_overrides(add_ovs, person)
            for ds, de, al_code in bundles:
                status = 'Confirmed' if al_code == 'AL_APPROVED' else 'Pending'

                # Check for 5-week span warnings
                if status == 'Pending' and ds <= five_week_end:
                    warnings.append(
                        f"{person}: provisional AL on {ds.strftime('%d-%m-%Y')} "
                        f"is within the 5-week span"
                    )

                # Deduplicate — skip if an identical entry already exists
                dup = next((l for l in leave_list
                            if l.get('name')       == person
                            and l.get('date_start') == ds.isoformat()
                            and l.get('date_end')   == de.isoformat()
                            and l.get('status')     in {'Confirmed', 'Pending'}
                            ), None)
                if dup:
                    warnings.append(
                        f"{person}: AL {ds.strftime('%d-%m-%Y')}–{de.strftime('%d-%m-%Y')} "
                        f"already exists as {dup.get('status')} — no new entry created."
                    )
                    continue

                leave_list.append({
                    'id':          str(uuid.uuid4())[:8],
                    'name':        person,
                    'username':    email,
                    'on_behalf':   True,
                    'created_by':  session['username'],
                    'created_at':  now,
                    'date_start':  ds.isoformat(),
                    'date_end':    de.isoformat(),
                    'leave_type':  'Annual Leave',
                    'status':      status,
                    'actioned_by': session['username'] if status == 'Confirmed' else None,
                    'actioned_at': now if status == 'Confirmed' else None,
                    'history': (
                        [{'status': 'Pending',   'by': session['username'], 'at': now},
                         {'status': 'Confirmed', 'by': session['username'], 'at': now}]
                        if status == 'Confirmed'
                        else [{'status': 'Pending', 'by': session['username'], 'at': now}]
                    ),
                })
                al_created += 1

        # ── Handle AL removals — find matching leave entries ──────────────
        for ov in remove_ovs:
            ov_date = date.fromisoformat(ov['date'])
            # Find leave entries that cover this date for this person.
            # We must also check the expanded (flanking-OFF) range because
            # _build_leave_map extends confirmed AL to adjacent OFF days —
            # so a user may have placed a "None" override on a flanking day
            # that isn't within the stored date_start/date_end.
            for entry in leave_list:
                if entry.get('name') != person:
                    continue
                try:
                    entry_ds = date.fromisoformat(entry['date_start'])
                    entry_de = date.fromisoformat(entry['date_end'])
                except (KeyError, ValueError):
                    continue
                # Check both the stored range and the expanded range
                current_status = entry.get('status', '')
                if current_status not in AL_APPROVED_STATUSES | AL_PENDING_STATUSES:
                    continue
                exp_ds, exp_de = _flanking_off_range(person, entry_ds, entry_de)
                if not (exp_ds <= ov_date <= exp_de):
                    continue
                if current_status in AL_CLEAR_STATUSES:
                    continue  # already cleared
                new_status = ('Withdrawn' if current_status in AL_APPROVED_STATUSES
                              else 'Rejected')
                allowed = VALID_TRANSITIONS.get(current_status, set())
                if new_status not in allowed:
                    continue
                entry['status']      = new_status
                entry['actioned_by'] = session['username']
                entry['actioned_at'] = now
                if 'history' not in entry:
                    entry['history'] = []
                entry['history'].append({'status': new_status,
                                         'by': session['username'], 'at': now})

            # Also write a published_override for this specific cell so
            # _resolve_shift (which checks override_map first) returns the
            # correct base shift immediately after publish, without waiting
            # for the leave_map to be rebuilt.
            base = _base_shift(person, ov_date)
            published_overrides = [
                p for p in published_overrides
                if not (p['person'] == person and p['date'] == ov['date'])
            ]
            published_overrides.append({
                'id':           str(uuid.uuid4())[:8],
                'person':       person,
                'date':         ov['date'],
                'shift':        base,
                'note':         ov.get('note'),
                'type':         'al_remove',
                'published_by': session['username'],
                'published_at': now,
            })

    # ── Process shift_change / weekend_toggle / coverage_swap ────────────
    for ov in other_overrides:
        published_overrides = [
            p for p in published_overrides
            if not (p['person'] == ov['person'] and p['date'] == ov['date'])
        ]
        published_overrides.append({
            'id':           str(uuid.uuid4())[:8],
            'person':       ov['person'],
            'date':         ov['date'],
            'shift':        ov['shift'],
            'note':         ov.get('note'),
            'type':         ov.get('type', 'shift_change'),
            'published_by': session['username'],
            'published_at': now,
        })
        shift_applied += 1

    _save_json(LEAVE_FILE, leave_list)
    _save_json(PUBLISHED_OVERRIDES_FILE, published_overrides)
    _save_draft_overrides([])
    _clear_draft_lock()

    return jsonify({
        'ok':           True,
        'published':    len(overrides),
        'al_created':   al_created,
        'shift_applied': shift_applied,
        'warnings':     warnings,
    })


# ════════════════════════════════════════════════════════════════════════════
#  WEEKEND SWAP
# ════════════════════════════════════════════════════════════════════════════

# All 3 engineering patterns that can appear in a Fri–Mon coverage window.
# Each entry: (before_sequence, after_sequence)
# 10-cell window = Wed, Thu, Fri, Sat, Sun, Mon, Tue, Wed, Thu, Fri+7
# Indices:         0    1    2    3    4    5    6    7    8    9

WEEKEND_SWAP_PATTERNS = [
    # Pattern 1 — Case B: covering eng had Fri(A)+Mon(A), Sat+Sun OFF
    (
        ('0900-1800','0900-1800','0900-1800','OFF','OFF','0900-1800','1000-2000','OFF','OFF','1000-2000'),
        ('0900-1800','OFF','0900-1800','0900-1800','0900-1800','1000-2000','OFF','OFF','OFF','1000-2000'),
    ),
    # Pattern 2 — Case A: covering eng was OFF all Fri–Mon
    (
        ('1000-2000','1000-2000','OFF','OFF','OFF','OFF','0900-1800','0900-1800','0900-1800','0900-1800'),
        ('OFF','OFF','1000-2000','1000-2000','0900-1800','0900-1800','OFF','0900-1800','0900-1800','OFF'),
    ),
]

# Coverage note indices (0-based within the 10-cell window) = Fri, Sat, Sun, Mon = 2,3,4,5
COVERAGE_NOTE_INDICES = {2, 3, 4, 5}


from typing import Optional

def _infer_absent_engineer(fri_date: date, leave_map: dict) -> Optional[str]:
    """Find which engineer has AL on Sat+Sun of the given weekend."""
    sat = fri_date + timedelta(days=1)
    sun = fri_date + timedelta(days=2)
    for name in ENGINEERING_OFFSETS:
        sat_leave = leave_map.get((name, sat))
        sun_leave = leave_map.get((name, sun))
        if (sat_leave and sat_leave['status'] in AL_APPROVED_STATUSES | AL_PENDING_STATUSES
                and sun_leave and sun_leave['status'] in AL_APPROVED_STATUSES | AL_PENDING_STATUSES):
            return name
    return None


def _get_window_shifts(person: str, fri_date: date,
                       leave_map: dict, override_map: dict) -> tuple:
    """Return the 10-cell window (Wed–Fri+7) as a tuple of shift strings."""
    wed = fri_date - timedelta(days=2)
    return tuple(
        _resolve_shift(person, wed + timedelta(days=i), leave_map, override_map)
        for i in range(10)
    )


def _match_weekend_pattern(window: tuple):
    """Return (pattern_idx, direction) where direction is 'swap' or 'revert',
    or None if no pattern matches."""
    for idx, (before, after) in enumerate(WEEKEND_SWAP_PATTERNS):
        if window == before:
            return idx, 'swap'
        if window == after:
            return idx, 'revert'
    return None


@rota_bp.route('/rota/draft/weekend-swap', methods=['PUT'])
@require_auth
def rota_draft_weekend_swap():
    """Apply or revert a weekend coverage swap for one engineering member.
    Requires draft lock. Expects {person, fri_date} in body."""
    err = _require_management()
    if err: return err
    err = _require_draft_lock_held_by_me()
    if err: return err

    session = request.session
    data    = request.get_json(silent=True) or {}
    person  = data.get('person', '').strip()
    fri_s   = data.get('fri_date', '').strip()

    if not person or not fri_s:
        return jsonify({'ok': False, 'error': 'person and fri_date required'}), 400
    if person not in ENGINEERING_OFFSETS:
        return jsonify({'ok': False, 'error': f'{person} is not an engineer'}), 400
    try:
        fri_date = date.fromisoformat(fri_s)
    except ValueError:
        return jsonify({'ok': False, 'error': 'Invalid date format'}), 400
    if fri_date.weekday() != 4:  # 4 = Friday
        return jsonify({'ok': False, 'error': 'fri_date must be a Friday'}), 400

    # Load all state
    overrides = _load_draft_overrides()
    leave_list = _load_json(LEAVE_FILE)
    if not isinstance(leave_list, list):
        leave_list = []
    published_overrides = _load_json(PUBLISHED_OVERRIDES_FILE)
    if not isinstance(published_overrides, list):
        published_overrides = []

    leave_map     = _build_leave_map(leave_list)
    published_map = _build_override_map(published_overrides)
    draft_map     = _build_override_map(overrides)
    combined_map  = {**published_map, **draft_map}

    # Read current 10-cell window
    window = _get_window_shifts(person, fri_date, leave_map, combined_map)
    match  = _match_weekend_pattern(window)

    if match is None:
        return jsonify({
            'ok':    False,
            'error': 'No matching weekend swap pattern found for this window.',
            'window': list(window),
        }), 400

    pattern_idx, direction = match
    before_seq, after_seq  = WEEKEND_SWAP_PATTERNS[pattern_idx]
    target_seq = after_seq if direction == 'swap' else before_seq

    # Infer absent engineer for note
    absent = _infer_absent_engineer(fri_date, leave_map)
    if absent:
        note_text = f'Covering for {absent} weekend absence'
    else:
        note_text = 'Covering weekend absence'

    # Apply the 10 overrides
    wed = fri_date - timedelta(days=2)
    now = _now_iso()

    notes = _load_notes()

    for i, new_shift in enumerate(target_seq):
        cell_date = wed + timedelta(days=i)
        date_s    = cell_date.isoformat()
        # Note only on Fri–Mon (indices 2–5) when swapping, not reverting
        if direction == 'swap' and i in COVERAGE_NOTE_INDICES:
            original = _base_shift(person, cell_date)
            cell_note = f'{note_text} | Original shift: {original}'
        else:
            cell_note = None

        # Remove any existing draft override for this cell
        overrides = [o for o in overrides
                     if not (o['person'] == person and o['date'] == date_s)]

        previous_shift = _resolve_shift(person, cell_date, leave_map, published_map)

        overrides.append({
            'id':             str(uuid.uuid4())[:8],
            'person':         person,
            'date':           date_s,
            'shift':          new_shift,
            'previous_shift': previous_shift,
            'note':           cell_note,
            'type':           'weekend_swap',
            'created_by':     session['username'],
            'created_at':     now,
        })

        # Sync the same note into the cell-notes store so it renders through
        # the normal note pipeline (_build_note_map), same as any manually
        # added note. Clears it on revert.
        if i in COVERAGE_NOTE_INDICES:
            notes = [n for n in notes
                     if not (n['person'] == person and n['date'] == date_s)]
            if cell_note:
                notes.append({
                    'id':         str(uuid.uuid4())[:8],
                    'person':     person,
                    'date':       date_s,
                    'note':       cell_note,
                    'created_by': session['username'],
                    'created_at': now,
                })

    _save_draft_overrides(overrides)
    _save_notes(notes)
    return jsonify({
        'ok':        True,
        'direction': direction,
        'pattern':   pattern_idx,
        'absent':    absent,
        'overrides': overrides,
        'window_before': list(window),
        'window_after':  list(target_seq),
    })


# ════════════════════════════════════════════════════════════════════════════
#  HOURS COMPUTATION ENGINE
# ════════════════════════════════════════════════════════════════════════════

# Net night minutes (22:00–07:00 window, after 1h lunch break deducted
# from cheapest portion first) per shift code.
# Lunch break (60 min) only applied to shifts ≥ 6 h (all named shifts qualify).
# Priority for deduction: daytime first, then night, then PH.
#
# Derivation per shift:
#   0700-1800  11h all day  → break from day → 0 night min
#   0900-1800   9h all day  → break from day → 0 night min
#   0900-2000  11h all day  → break from day → 0 night min
#   0930-1800  8.5h all day → break from day → 0 night min
#   0900-1730  8.5h all day → break from day → 0 night min
#   0800-1630  8.5h all day → break from day → 0 night min
#   1000-2020  10h all day  → break from day → 0 night min  (note: 1000-2020 typo; real shift is 1000-2000)
#   1000-2000  10h all day  → break from day → 0 night min
#   1300-0000  11h: 9h day (13-22) + 2h night (22-00) → break from day → 2h night = 120 min
#   1500-0200  11h: 7h day (15-22) + 4h night (22-02) → break from day → 4h night = 240 min
#   2100-0700  10h: 1h day (21-22) + 9h night (22-07) → break fully from day → 9h night = 540 min

SHIFT_NIGHT_MINUTES: dict[str, int] = {
    '0700-1800': 0,
    '0900-1800': 0,
    '0900-2000': 0,
    '0930-1800': 0,
    '0900-1730': 0,
    '0800-1630': 0,
    '1000-2000': 0,
    '1300-0000': 120,   # 2h
    '1500-0200': 240,   # 4h
    '2100-0700': 540,   # 9h
    'OFF':       0,
}

# Total shift duration in minutes (raw, before lunch break) per code
SHIFT_TOTAL_MINUTES: dict[str, int] = {
    '0700-1800': 660,
    '0900-1800': 540,
    '0900-2000': 660,
    '0930-1800': 510,
    '0900-1730': 510,
    '0800-1630': 510,
    '1000-2000': 600,
    '1300-0000': 660,
    '1500-0200': 660,
    '2100-0700': 600,
    'OFF':       0,
}

LUNCH_BREAK_MINUTES = 60
LUNCH_BREAK_THRESHOLD_MINUTES = 360  # 6h

def _net_minutes(shift_code: str) -> tuple[int, int]:
    """Return (net_daytime_minutes, net_night_minutes) after lunch break.
    Break is deducted from daytime first. Returns (0,0) for OFF/unknown.
    Named shifts use the fixed tables; other HHMM-HHMM codes (e.g. ad-hoc
    Engineering overrides) fall back to generic time parsing."""
    total, night, _, _ = _shift_minutes_lookup(shift_code)
    if total == 0:
        return (0, 0)
    day = total - night
    break_min = LUNCH_BREAK_MINUTES if total >= LUNCH_BREAK_THRESHOLD_MINUTES else 0
    # Deduct from daytime first
    day_after = max(0, day - break_min)
    remaining_break = max(0, break_min - day)
    night_after = max(0, night - remaining_break)
    return (day_after, night_after)


# Same-day night minutes (22:00–00:00 of the shift's own calendar day)
# after lunch break. Used for PH NH when the shift is ON the PH date.
SHIFT_SAME_DAY_NIGHT_MINUTES: dict[str, int] = {
    '0700-1800': 0,
    '0900-1800': 0,
    '0900-2000': 0,
    '0930-1800': 0,
    '0900-1730': 0,
    '0800-1630': 0,
    '1000-2000': 0,
    '1300-0000': 120,   # 22:00–00:00 = 2h
    '1500-0200': 120,   # 22:00–00:00 = 2h
    '2100-0700': 120,   # 22:00–00:00 = 2h (lunch consumed daytime 21-22)
    'OFF':       0,
}

# Spillover night minutes (00:00–shift_end on the NEXT calendar day)
SHIFT_SPILLOVER_NIGHT_MINUTES: dict[str, int] = {
    '0700-1800': 0,
    '0900-1800': 0,
    '0900-2000': 0,
    '0930-1800': 0,
    '0900-1730': 0,
    '0800-1630': 0,
    '1000-2000': 0,
    '1300-0000': 0,     # ends exactly at midnight, no spillover
    '1500-0200': 120,   # 00:00–02:00 = 2h
    '2100-0700': 420,   # 00:00–07:00 = 7h
    'OFF':       0,
}


# ── Generic fallback for shift codes NOT in the fixed tables above ─────────
# Named rotation shifts (Management/Engineering-base/Specialist) are always
# resolved from the dicts above — this is only reached for arbitrary
# HHMM-HHMM override codes (e.g. ad-hoc Engineering shift changes) that
# don't match any known named shift. Night window: 22:00-07:00.
def _parse_raw_shift_minutes(code: str):
    """Return (total_min, night_min, same_day_night_min, spillover_night_min)
    for an arbitrary 'HHMM-HHMM' code, or None if unparseable."""
    if not code or code == 'OFF':
        return (0, 0, 0, 0)
    try:
        s_str, e_str = code.split('-')
        sh, sm = int(s_str[:2]), int(s_str[2:])
        eh, em = int(e_str[:2]), int(e_str[2:])
    except (ValueError, IndexError):
        return None

    base = datetime.datetime(2000, 1, 1)
    start = base.replace(hour=sh, minute=sm)
    end   = base.replace(hour=eh, minute=em)
    if end <= start:
        end += datetime.timedelta(days=1)
    total = int((end - start).total_seconds() // 60)

    night_start = base.replace(hour=22, minute=0)
    night_end   = base + datetime.timedelta(days=1, hours=7)
    ov_s, ov_e  = max(start, night_start), min(end, night_end)
    night = max(0, int((ov_e - ov_s).total_seconds() // 60)) if ov_s < ov_e else 0

    same_day_start = base.replace(hour=22, minute=0)
    same_day_end   = base + datetime.timedelta(days=1)
    sd_s, sd_e     = max(start, same_day_start), min(end, same_day_end)
    same_day = max(0, int((sd_e - sd_s).total_seconds() // 60)) if sd_s < sd_e else 0

    spill_start = base + datetime.timedelta(days=1)
    spill = max(0, int((end - spill_start).total_seconds() // 60)) if end > spill_start else 0

    return (total, night, same_day, spill)


def _shift_minutes_lookup(code: str):
    """Return (total, night, same_day_night, spillover_night) for any shift
    code: named shifts use the fixed tables (unchanged behaviour); anything
    else falls back to generic HHMM-HHMM parsing."""
    if code in SHIFT_TOTAL_MINUTES:
        return (SHIFT_TOTAL_MINUTES[code],
                SHIFT_NIGHT_MINUTES.get(code, 0),
                SHIFT_SAME_DAY_NIGHT_MINUTES.get(code, 0),
                SHIFT_SPILLOVER_NIGHT_MINUTES.get(code, 0))
    parsed = _parse_raw_shift_minutes(code)
    return parsed if parsed is not None else (0, 0, 0, 0)


def _effective_shift_for_hours(name: str, d: date,
                                leave_map: dict, override_map: dict) -> str:
    """Return the shift code to use for NH accounting.
    For any override: use max(NH(new), NH(original)) to protect entitlement.
    ABSENT and leave states → 'OFF' (0h)."""
    def _clean(s):
        if not s or s in ('OFF', 'PARENTAL', 'MARITAL'): return 'OFF'
        if s.startswith('AL_') or s.startswith('ABSENT'): return 'OFF'
        if '|' in s:
            pfx = s.split('|', 1)[0]
            if pfx in ('AL_APPROVED', 'AL_PENDING', 'ABSENT'): return 'OFF'
            return s.split('|', 1)[1]
        return s

    if override_map:
        ov = override_map.get((name, d))
        if ov is not None:
            raw = ov.get('shift', 'OFF')
            # Absence/leave overrides never generate NH entitlement,
            # regardless of what was originally rostered.
            if raw.startswith('ABSENT') or raw.startswith('AL_'):
                return 'OFF'
            new_clean  = _clean(raw)
            orig_clean = _clean(_base_shift(name, d))
            _, new_nh  = _net_minutes(new_clean)
            _, orig_nh = _net_minutes(orig_clean)
            return new_clean if new_nh >= orig_nh else orig_clean

    resolved = _resolve_shift(name, d, leave_map, override_map)
    return _clean(resolved)

# ── POT helpers ───────────────────────────────────────────────────────────

# ── AL Allowance helpers ────────────────────────────────────────────────

ABSENCE_ZERO_DAYS = 3   # awarded when zero ABSENT records exist in N-1

def _load_al_file() -> dict:
    data = _load_json(AL_ALLOWANCE_FILE)
    if not isinstance(data, dict):
        data = {}
    data.setdefault('members', {})
    data.setdefault('yearly', {})
    return data

def _save_al_file(data: dict) -> None:
    _save_json(AL_ALLOWANCE_FILE, data)

def _al_member_join_date(al: dict, name: str):
    m = al['members'].get(name)
    if not m or not m.get('join_date'):
        return None
    try:
        return date.fromisoformat(m['join_date'])
    except ValueError:
        return None

def _scan_absent_days(person: str, year: int) -> int:
    """Count ABSENT-status shift_change overrides for a person across the
    given calendar year, scanned from published_overrides.json."""
    published = _load_json(PUBLISHED_OVERRIDES_FILE)
    if not isinstance(published, list):
        published = []
    count = 0
    for o in published:
        if o.get('person') != person:
            continue
        if o.get('type') != 'shift_change':
            continue
        shift = o.get('shift') or ''
        if not shift.startswith('ABSENT'):
            continue
        try:
            d = date.fromisoformat(o['date'])
        except (KeyError, ValueError):
            continue
        if d.year == year:
            count += 1
    return count

def _blank_field():
    return {'computed_days': None, 'final_days': None,
            'auto_or_manual': 'auto', 'override_reason': None}

def _evaluate_member_year(al: dict, name: str, year: int) -> dict:
    """Compute default values for base_allowance, mhd, absence for one
    member/year. Does NOT overwrite fields already set to 'manual' by a
    prior save — only fills computed_days and, for fields still 'auto',
    final_days too."""
    join = _al_member_join_date(al, name)
    yr_block  = al['yearly'].setdefault(str(year), {})
    yr_block.setdefault('mhd_default_days', None)
    yr_members = yr_block.setdefault('members', {})
    existing = yr_members.get(name, {})

    base_f    = existing.get('base_allowance', _blank_field())
    mhd_f     = existing.get('mhd', _blank_field())
    abs_f     = existing.get('absence', {**_blank_field(),
                                          'flagged_for_review': False})

    joined_before_or_during_prev_year = join is not None and join.year <= year - 1
    joined_this_year                  = join is not None and join.year >= year

    # ── base_allowance ──────────────────────────────────────────────────
    if base_f['auto_or_manual'] == 'auto':
        if joined_before_or_during_prev_year:
            base_f['computed_days'] = 22
            base_f['final_days']    = 22
        else:
            base_f['computed_days'] = None
            base_f['final_days']    = base_f.get('final_days')  # leave for manual entry

    # ── mhd ─────────────────────────────────────────────────────────────
    mhd_default = yr_block.get('mhd_default_days')
    if mhd_f['auto_or_manual'] == 'auto':
        if joined_before_or_during_prev_year and mhd_default is not None:
            mhd_f['computed_days'] = mhd_default
            mhd_f['final_days']    = mhd_default
        else:
            mhd_f['computed_days'] = None
            mhd_f['final_days']    = mhd_f.get('final_days')

    # ── absence ─────────────────────────────────────────────────────────
    if abs_f['auto_or_manual'] == 'auto':
        if joined_this_year:
            abs_f['computed_days']      = 0
            abs_f['final_days']         = 0
            abs_f['flagged_for_review'] = False
        elif joined_before_or_during_prev_year:
            absent_count = _scan_absent_days(name, year - 1)
            if absent_count == 0:
                abs_f['computed_days']      = ABSENCE_ZERO_DAYS
                abs_f['final_days']         = ABSENCE_ZERO_DAYS
                abs_f['flagged_for_review'] = False
            else:
                abs_f['computed_days']      = None
                abs_f['final_days']         = abs_f.get('final_days')
                abs_f['flagged_for_review'] = True
        else:
            # no join_date on file at all — can't evaluate
            abs_f['computed_days']      = None
            abs_f['final_days']         = abs_f.get('final_days')
            abs_f['flagged_for_review'] = False

    result = {
        'evaluated':        True,
        'base_allowance':   base_f,
        'mhd':              mhd_f,
        'absence':          abs_f,
        'misc_entries':     existing.get('misc_entries', []),
        'carry_over_hours': existing.get('carry_over_hours', 0),
        'carry_over_locked': existing.get('carry_over_locked', False),
    }
    yr_members[name] = result
    return result

def _ph_al_giveback_hours(name: str, year: int,
                          leave_map: dict) -> tuple[float, list]:
    """Live-computed PH-on-AL hours given back for a calendar year.
    Returns (total_hours, list_of_dates_str). Never persisted."""
    total_min = 0
    dates = []
    for ph_date in PUBLIC_HOLIDAYS:
        if ph_date.year != year:
            continue
        leave = leave_map.get((name, ph_date))
        if not leave or leave['status'] not in AL_APPROVED_STATUSES | AL_PENDING_STATUSES:
            continue
        base = _base_shift(name, ph_date)
        if base == 'OFF':
            continue  # wasn't expected to work anyway — nothing to give back
        total_min_shift, _ = _net_minutes(base)
        _, night_min       = _net_minutes(base)
        shift_total_min     = total_min_shift + night_min  # net day + net night
        total_min += shift_total_min
        dates.append(ph_date.isoformat())
    return round(total_min / 60, 2), sorted(dates)

def _compute_al_used_hours(name: str, year: int,
                           leave_list: list) -> tuple[float, float]:
    """Return (confirmed_used_hours, pending_used_hours) for AL leave
    entries in the given year, using actual scheduled shift duration
    per day (not flat 8h)."""
    confirmed_min = 0
    pending_min   = 0
    for r in leave_list:
        if r.get('name') != name or r.get('leave_type') != 'Annual Leave':
            continue
        status = r.get('status')
        if status not in (AL_APPROVED_STATUSES | AL_PENDING_STATUSES):
            continue
        try:
            ds = date.fromisoformat(r['date_start'])
            de = date.fromisoformat(r['date_end'])
        except (KeyError, ValueError):
            continue
        d = ds
        while d <= de:
            if d.year == year:
                base = _base_shift(name, d)
                if base != 'OFF':
                    day_min, night_min = _net_minutes(base)
                    mins = day_min + night_min
                    if status in AL_APPROVED_STATUSES:
                        confirmed_min += mins
                    else:
                        pending_min += mins
            d += timedelta(days=1)
    return round(confirmed_min / 60, 2), round(pending_min / 60, 2)

def _compute_al_balance(al: dict, name: str, year: int,
                        leave_list: list, leave_map: dict) -> dict:
    entry = _evaluate_member_year(al, name, year)

    base_h  = (entry['base_allowance']['final_days'] or 0) * 8
    mhd_h   = (entry['mhd']['final_days'] or 0) * 8
    abs_h   = (entry['absence']['final_days'] or 0) * 8
    misc_h  = sum(e.get('hours', 0) for e in entry.get('misc_entries', []))
    carry_h = entry.get('carry_over_hours', 0)
    ph_h, ph_dates = _ph_al_giveback_hours(name, year, leave_map)

    total_allowance = base_h + mhd_h + abs_h + misc_h + carry_h + ph_h
    confirmed_used, pending_used = _compute_al_used_hours(name, year, leave_list)

    return {
        'name':               name,
        'year':               year,
        'base_allowance':     entry['base_allowance'],
        'mhd':                entry['mhd'],
        'absence':            entry['absence'],
        'misc_entries':       entry.get('misc_entries', []),
        'misc_hours':         misc_h,
        'carry_over_hours':   carry_h,
        'carry_over_locked':  entry.get('carry_over_locked', False),
        'ph_al_giveback_hours': ph_h,
        'ph_al_giveback_dates': ph_dates,
        'total_allowance_hours':      total_allowance,
        'confirmed_used_hours':       confirmed_used,
        'pending_used_hours':         pending_used,
        'remaining_confirmed_hours':  round(total_allowance - confirmed_used, 2),
        'remaining_with_pending_hours': round(total_allowance - confirmed_used - pending_used, 2),
    }

def _load_pot() -> list:
    data = _load_json(HOURS_POT_FILE)
    return data if isinstance(data, list) else []

def _save_pot(records: list) -> None:
    _save_json(HOURS_POT_FILE, records)

from typing import Union

def _active_pot_for(team: str, month: str) -> Union[dict, None]:
    """Return the active POT record for team+month, or None."""
    return next((r for r in _load_pot()
                 if r['team'] == team and r['month'] == month
                 and r['status'] == 'active'), None)

def _check_hr_config_consistency() -> list:
    """Compare person_directory.json HR team membership against the live
    users API. Returns a list of warning strings — empty if consistent."""
    users = _load_json(USERS_FILE)
    if not isinstance(users, dict):
        users = {}

    email_to_label = {
        email: _RID_TO_LABEL[rid]
        for email, rid in _email_to_employee_id().items()
        if rid in _RID_TO_LABEL
    }
    label_to_email = {v: k for k, v in email_to_label.items()}

    valid_roles = STAFF_ROLES | {'admin'}
    active_emails = {email for email, info in users.items()
                     if isinstance(info, dict) and info.get('role') in valid_roles}

    hr_teams = _hr_teams()
    all_hr_names = {name for members in hr_teams.values() for name in members}

    warnings = []

    # Type A: in an HR team but not resolvable to an active users.json entry
    for name in sorted(all_hr_names):
        email = label_to_email.get(name)
        if not email:
            warnings.append(
                f"Type A — '{name}' has an hr_team set in person_directory.json "
                f"but no matching employee_id found via the users API. Hours "
                f"will be computed but not linked to any login account."
            )
        elif email not in active_emails:
            warnings.append(
                f"Type A — '{name}' ({email}) is in an HR team but has no "
                f"active entry in users.json. They may have left — check "
                f"person_directory.json."
            )

    # Type B: active engineer/specialist in users.json not in any HR team
    for email, info in users.items():
        if not isinstance(info, dict):
            continue
        if info.get('role') not in STAFF_ROLES:
            continue
        rota_name = email_to_label.get(email)
        if not rota_name:
            continue  # user exists but has no directory entry — separate issue
        if rota_name not in all_hr_names:
            warnings.append(
                f"Type B — '{rota_name}' ({email}) has role "
                f"'{info.get('role')}' in users.json but has no hr_team set "
                f"in person_directory.json. Their hours will never be "
                f"computed or exported to HR."
            )

    return warnings


def _compute_hours(date_from: date, date_to: date,
                   names: list[str],
                   leave_map: dict, override_map: dict) -> dict:
    """Compute night and PH hours for each name over the date range.

    PH hours split by calendar day boundary:
    - Shift ON a PH: earns PH NH for 22:00–00:00 (same-day night only).
      PH daytime = net daytime of that shift.
    - Shift on DAY BEFORE a PH: spillover (00:00–shift_end) counts as PH NH.
    - Regular NH: total NH entitlement minus any portions counted as PH NH.

    Returns dict: name → {night_h, ph_day_h, ph_night_h, ph_dates}
    """
    results = {n: {'night_min': 0, 'ph_day_min': 0,
                   'ph_night_min': 0, 'ph_dates': set()} for n in names}

    def _clean_actual(s):
        if not s or s in ('OFF', 'PARENTAL', 'MARITAL'): return 'OFF'
        if s.startswith('AL_') or s.startswith('ABSENT'): return 'OFF'
        if '|' in s:
            pfx = s.split('|', 1)[0]
            if pfx in ('AL_APPROVED', 'AL_PENDING', 'ABSENT'): return 'OFF'
            return s.split('|', 1)[1]
        return s

    # Start one day before date_from to catch spillover into first day
    d = date_from - timedelta(days=1)

    while d <= date_to:
        is_ph      = d in PUBLIC_HOLIDAYS
        next_is_ph = (d + timedelta(days=1)) in PUBLIC_HOLIDAYS
        in_range   = d >= date_from

        for name in names:
            nh_shift     = _effective_shift_for_hours(name, d, leave_map, override_map)
            actual_shift = _clean_actual(_resolve_shift(name, d, leave_map, override_map))

            if actual_shift == 'OFF' and nh_shift == 'OFF':
                continue

            _, night_min_nh   = _net_minutes(nh_shift)
            day_min_actual, _ = _net_minutes(actual_shift)
            _, _, same_day_night, spillover_night = _shift_minutes_lookup(actual_shift)

            # ── PH: same-day contribution ─────────────────────────────────
            if is_ph and in_range and actual_shift != 'OFF':
                results[name]['ph_day_min']  += day_min_actual
                results[name]['ph_night_min'] += same_day_night
                if day_min_actual + same_day_night > 0:
                    results[name]['ph_dates'].add(d)

            # ── PH: spillover from previous day into next (PH) day ────────
            if next_is_ph and spillover_night > 0 and actual_shift != 'OFF':
                next_d = d + timedelta(days=1)
                if next_d >= date_from:
                    results[name]['ph_night_min'] += spillover_night
                    results[name]['ph_dates'].add(next_d)

            # ── Regular NH — independent of PH status ───────────────────────
            # NH and PH-NH are separate, additive entitlements: a night hour
            # worked on (or spilling into) a public holiday still counts as
            # a regular NH AND as a PH-NH. No exclusion, no deduction.
            if in_range:
                results[name]['night_min'] += night_min_nh

        d += timedelta(days=1)

    out = {}
    for name, r in results.items():
        ph_date_strs = sorted(dd.strftime('%d %B').lstrip('0') for dd in r['ph_dates'])
        out[name] = {
            'night_h':    round(r['night_min'] / 60, 2),
            'ph_day_h':   round(r['ph_day_min'] / 60, 2),
            'ph_night_h': round(r['ph_night_min'] / 60, 2),
            'ph_dates':   ph_date_strs,
        }
    return out


# ── Hours routes ───────────────────────────────────────────────────────────

@rota_bp.route('/rota/hours', methods=['GET'])
@require_auth
def rota_hours_get():
    """Compute night + PH hours for a date range.
    Management: all members. Staff: own row only."""
    session   = request.session
    rota_role = _get_rota_role(session)
    username  = session['username']

    try:
        date_from = date.fromisoformat(
            request.args.get('from', date.today().replace(day=1).isoformat()))
        date_to   = date.fromisoformat(
            request.args.get('to', date.today().isoformat()))
    except ValueError:
        return jsonify({'ok': False, 'error': 'Invalid date format, use YYYY-MM-DD'}), 400

    leave_list = _load_json(LEAVE_FILE)
    if not isinstance(leave_list, list):
        leave_list = []
    published_overrides = _load_json(PUBLISHED_OVERRIDES_FILE)
    if not isinstance(published_overrides, list):
        published_overrides = []

    leave_map    = _build_leave_map(leave_list)
    override_map = _build_override_map(published_overrides)

    all_names = (list(MANAGEMENT_SHIFTS) +
                 list(ENGINEERING_OFFSETS) +
                 list(SPECIALIST_OFFSETS))

    if rota_role != 'management':
        # Staff: only own name
        my_name = _rota_display_name(username)
        names = [my_name] if my_name in all_names else []
    else:
        names = all_names

    hours = _compute_hours(date_from, date_to, names, leave_map, override_map)

    # Annotate with team
    hr_teams = _hr_teams()
    name_to_team = {}
    for team, members in hr_teams.items():
        for m in members:
            name_to_team[m] = team
    # Also add rota team for display grouping
    rota_team_map = {}
    for n in MANAGEMENT_SHIFTS:
        rota_team_map[n] = 'Management'
    for n in ENGINEERING_OFFSETS:
        rota_team_map[n] = 'Engineering'
    for n in SPECIALIST_OFFSETS:
        rota_team_map[n] = 'Specialists'

    result = {}
    for name, h in hours.items():
        result[name] = {
            **h,
            'rota_team': rota_team_map.get(name, 'Unknown'),
            'hr_team':   name_to_team.get(name),
        }

    return jsonify({'ok': True, 'from': date_from.isoformat(),
                    'to': date_to.isoformat(), 'hours': result})


@rota_bp.route('/rota/hours/export', methods=['GET'])
@require_auth
def rota_hours_export():
    """Generate HR Excel sheet for a team and month. Management only."""
    if _get_rota_role(request.session) != 'management':
        return jsonify({'ok': False, 'error': 'Not authorised'}), 403

    team_param = request.args.get('team', '').upper()
    month_param = request.args.get('month', '')  # YYYY-MM

    if team_param not in ('SOE', 'SOS'):
        return jsonify({'ok': False, 'error': 'team must be SOE or SOS'}), 400
    try:
        year, month = [int(x) for x in month_param.split('-')]
        date_from = date(year, month, 1)
        # Last day of month
        if month == 12:
            date_to = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            date_to = date(year, month + 1, 1) - timedelta(days=1)
    except (ValueError, AttributeError):
        return jsonify({'ok': False, 'error': 'month must be YYYY-MM'}), 400

    leave_list = _load_json(LEAVE_FILE)
    if not isinstance(leave_list, list):
        leave_list = []
    published_overrides = _load_json(PUBLISHED_OVERRIDES_FILE)
    if not isinstance(published_overrides, list):
        published_overrides = []

    # ── Serve from POT — never recompute for export ───────────────────────
    pot = _active_pot_for(team_param, month_param)
    if not pot:
        return jsonify({
            'ok':    False,
            'error': (f'No committed Point of Truth exists for {team_param} '
                      f'{month_param}. Compute, review, and commit via the '
                      f'Night & PH Hours tab before exporting.'),
            'needs_commit': True,
        }), 400

    # Names, hours, and employee IDs all come straight from the POT
    # snapshot — captured at commit time, so export can never drift from
    # what was actually committed even if person_directory.json changes later.
    display_names = {e['name']: e.get('full_name', e['name']) for e in pot['entries']}
    members = [e['name'] for e in pot['entries']]
    hours   = {
        e['name']: {
            'night_h':    e['night_h_final'],
            'ph_day_h':   e['ph_day_h_final'],
            'ph_night_h': e['ph_night_h_final'],
            'ph_dates':   e['ph_dates'],
        }
        for e in pot['entries']
    }
    emp_id_map = {e['name']: e.get('employeeID') for e in pot['entries']}

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'ok': False, 'error': 'openpyxl not available'}), 500

    wb = Workbook()
    ws = wb.active
    month_label = date_from.strftime('%B %Y')
    ws.title = f'{team_param} {month_label}'

    # ── Styles ────────────────────────────────────────────────────────────
    hdr_font  = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    hdr_fill  = PatternFill('solid', fgColor='000000')
    body_font = Font(name='Calibri', size=11, bold=True)
    num_font  = Font(name='Calibri', size=11, bold=True)
    note_font = Font(name='Calibri', size=11, color='595959')
    row_fill  = PatternFill('solid', fgColor='FFF2CC')   # yellow — numeric columns
    id_fill   = PatternFill('solid', fgColor='D9D9D9')   # gray  — ID/name columns
    center    = Alignment(horizontal='center', vertical='center')
    left      = Alignment(horizontal='left',   vertical='center')
    thin      = Side(style='thin', color='000000')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Header row (row 1) ────────────────────────────────────────────────
    headers = [
        'Employee ID', 'Employee',
        'Horas Noturnas\nNight hours (10pm-7 am)',
        'Feriado Diurnas\n| Public Holidays',
        'Feriado Noturnas\n| Night Holiday',
        'Holiday Date',
        'Horas Extra\n|Overtime hours',
        'Over time Date',
    ]
    col_widths = [11, 21, 14, 14, 14, 32, 14, 17]

    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
        cell           = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 60

    # ── Data rows ─────────────────────────────────────────────────────────
    for row_idx, name in enumerate(members, start=2):
        h      = hours.get(name, {'night_h': 0, 'ph_day_h': 0,
                                   'ph_night_h': 0, 'ph_dates': []})
        emp_id      = emp_id_map.get(name)
        full_name   = display_names.get(name, name)
        ph_str      = ', '.join(h['ph_dates']) if h['ph_dates'] else ''

        row_data = [
            emp_id if emp_id is not None else '',
            full_name,
            round(h['night_h'], 2),
            round(h['ph_day_h'], 2),
            round(h['ph_night_h'], 2),
            ph_str,
            0.00,
            '',
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell        = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill   = id_fill if col_idx <= 2 else row_fill
            cell.border = border
            if col_idx in (3, 4, 5, 7):   # numeric hour columns — bold
                cell.font          = num_font
                cell.number_format = '0.00'
                cell.alignment     = center
            elif col_idx == 1:             # Employee ID — centred, normal weight
                cell.font      = body_font
                cell.alignment = center
            else:
                cell.font      = body_font
                cell.alignment = center
        ws.row_dimensions[row_idx].height = 18

    # ── Footer notes (2 rows below last data row) ──────────────────────────
    note_row = len(members) + 3
    notes = [
        '* Kindly note if employee number is incorrect, person will not be paid.',
        '** Numbers need to have 2 decimal houses. No cell should be left empty. '
        'If there are no hours it should say 0.00',
    ]
    for i, note_text in enumerate(notes):
        cell      = ws.cell(row=note_row + i, column=1, value=note_text)
        cell.font = note_font
        ws.merge_cells(start_row=note_row + i, start_column=1,
                       end_row=note_row + i,   end_column=8)

    # ── Freeze panes below header ──────────────────────────────────────────
    ws.freeze_panes = 'A2'

    # ── Stream to response ─────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'{team_param}_NightHours_{date_from.strftime("%b%Y")}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

# ── PicaPonto (attendance) export ─────────────────────────────────────────

from typing import Optional

_PICAPONTO_ROLE_ORDER = [
    'Technical Operations Manager',
    'Streaming Ops Engineering Lead',
    'Streaming Ops Lead',
    'Streaming Ops Engineer',
    'Streaming Ops Specialist',
]


def _picaponto_job_role(email: str, info: dict) -> str:
    """Derive export Job Role from users.json fields for one user."""
    role        = info.get('role', '')
    team        = info.get('team', '')
    rota_status = info.get('rota_status', '')

    if role == 'admin' and team == 'na' and rota_status == 'active':
        return 'Technical Operations Manager'
    if role == 'admin' and team == 'soe':
        return 'Streaming Ops Engineering Lead'
    if role == 'admin' and team == 'sos':
        return 'Streaming Ops Lead'
    if role == 'engineer':
        return 'Streaming Ops Engineer'
    return 'Streaming Ops Specialist'


def _shift_to_times(shift: str):
    """Return (clock_in_time, clock_out_time) or (None, None) for OFF/leave."""
    import re as _re, datetime as _dt
    if not shift or shift == 'OFF':
        return None, None
    if shift.startswith('AL_') or shift.startswith('ABSENT') or shift in ('PARENTAL', 'MARITAL'):
        return None, None
    if '|' in shift:
        prefix, rest = shift.split('|', 1)
        if prefix in ('AL_APPROVED', 'AL_PENDING', 'ABSENT'):
            return None, None
        shift = rest
    m = _re.match(r'^(\d{2})(\d{2})-(\d{2})(\d{2})$', shift)
    if not m:
        return None, None
    sh, sm, eh, em = int(m[1]), int(m[2]), int(m[3]), int(m[4])
    return _dt.time(sh, sm), _dt.time(eh, em)


@rota_bp.route('/rota/picaponto-export', methods=['GET'])
@require_auth
def rota_picaponto_export():
    """Generate a PicaPonto-format attendance Excel for a month. Management only."""
    if _get_rota_role(request.session) != 'management':
        return jsonify({'ok': False, 'error': 'Not authorised'}), 403

    month_param = request.args.get('month', '')  # YYYY-MM
    try:
        year, month = [int(x) for x in month_param.split('-')]
        date_from = date(year, month, 1)
        date_to   = (date(year, 12, 31) if month == 12
                     else date(year, month + 1, 1) - timedelta(days=1))
    except (ValueError, AttributeError):
        return jsonify({'ok': False, 'error': 'month must be YYYY-MM'}), 400

    days_in_month = (date_to - date_from).days + 1

    leave_list = _load_json(LEAVE_FILE)
    if not isinstance(leave_list, list):
        leave_list = []
    published_overrides = _load_json(PUBLISHED_OVERRIDES_FILE)
    if not isinstance(published_overrides, list):
        published_overrides = []

    leave_map    = _build_leave_map(leave_list)
    override_map = _build_override_map(published_overrides)

    users_dict = _load_json(USERS_FILE)
    if not isinstance(users_dict, dict):
        users_dict = {}

    all_rota_names = (list(MANAGEMENT_SHIFTS) +
                      list(ENGINEERING_OFFSETS) +
                      list(SPECIALIST_OFFSETS))

    members = []
    for rota_name in all_rota_names:
        email = _email_for_rota_name(rota_name)
        u     = users_dict.get(email, {}) if email else {}
        members.append({
            'rota_name':   rota_name,
            'name':        _LABEL_TO_FULLNAME.get(rota_name, rota_name),
            'employee_id': _LABEL_TO_RID.get(rota_name, ''),
            'job_role':    _picaponto_job_role(email or '', u),
        })

    def _role_sort_key(m):
        try:
            return (_PICAPONTO_ROLE_ORDER.index(m['job_role']), m['name'])
        except ValueError:
            return (len(_PICAPONTO_ROLE_ORDER), m['name'])

    members.sort(key=_role_sort_key)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import datetime as _dt
    except ImportError:
        return jsonify({'ok': False, 'error': 'openpyxl not available'}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = 'Proposal'

    FONT_NAME  = 'Aptos Narrow'
    BLACK_FILL = PatternFill('solid', fgColor='000000')
    WHITE_FONT = Font(name=FONT_NAME, size=11, color='FFFFFF')
    BODY_FONT  = Font(name=FONT_NAME, size=11, color='000000')
    TIME_FMT   = 'h:mm'
    CENTER     = Alignment(horizontal='center', vertical='center')
    MED_SIDE   = Side(style='medium', color='000000')
    MED_LEFT   = Border(left=MED_SIDE)
    MED_RIGHT  = Border(right=MED_SIDE)

    for ci, w in enumerate([8.85, 20.28, 29.85, 12.85], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    day_sub_widths = [9.28, 13.0, 8.85, 9.28, 8.85]
    for day_idx in range(days_in_month):
        base_col = 5 + day_idx * 5
        for sub, w in enumerate(day_sub_widths):
            ws.column_dimensions[get_column_letter(base_col + sub)].width = w

    # Row 1: date headers
    for ci in range(1, 5):
        cell = ws.cell(1, ci)
        cell.fill = BLACK_FILL; cell.font = WHITE_FONT; cell.alignment = CENTER

    for day_idx in range(days_in_month):
        base_col = 5 + day_idx * 5
        d        = date_from + timedelta(days=day_idx)
        ws.merge_cells(start_row=1, start_column=base_col,
                       end_row=1,   end_column=base_col + 4)
        cell = ws.cell(1, base_col)
        cell.value         = _dt.datetime(d.year, d.month, d.day)
        cell.number_format = 'mm-dd-yy'
        cell.font          = WHITE_FONT
        cell.fill          = BLACK_FILL
        cell.alignment     = CENTER
        ws.cell(1, base_col).border     = MED_LEFT
        ws.cell(1, base_col + 4).border = MED_RIGHT

    # Row 2: sub-headers
    for ci, lbl in enumerate(['ID', 'Name', 'Job role', 'Department'], start=1):
        cell = ws.cell(2, ci)
        cell.value = lbl; cell.fill = BLACK_FILL
        cell.font  = WHITE_FONT; cell.alignment = CENTER
    ws.row_dimensions[2].height = 15.75

    for day_idx in range(days_in_month):
        base_col = 5 + day_idx * 5
        for sub in range(5):
            cell = ws.cell(2, base_col + sub)
            cell.fill = BLACK_FILL; cell.font = WHITE_FONT; cell.alignment = CENTER

    # Data rows
    for row_idx, m in enumerate(members, start=3):
        rota_name = m['rota_name']
        for ci, val in enumerate(
            [m['employee_id'], m['name'], m['job_role'], 'Video Ops'], start=1
        ):
            cell = ws.cell(row_idx, ci)
            cell.value = val; cell.font = BODY_FONT; cell.alignment = CENTER

        for day_idx in range(days_in_month):
            base_col = 5 + day_idx * 5
            d        = date_from + timedelta(days=day_idx)
            shift    = _resolve_shift(rota_name, d, leave_map, override_map)
            cin, cout = _shift_to_times(shift)

            cin_cell        = ws.cell(row_idx, base_col)
            cin_cell.font   = BODY_FONT
            cin_cell.alignment = CENTER
            cin_cell.border = MED_LEFT
            if cin is not None:
                cin_cell.value         = cin
                cin_cell.number_format = TIME_FMT

            cout_cell           = ws.cell(row_idx, base_col + 1)
            cout_cell.font      = BODY_FONT
            cout_cell.alignment = CENTER
            if cout is not None:
                cout_cell.value         = cout
                cout_cell.number_format = TIME_FMT

            ws.cell(row_idx, base_col + 4).border = MED_RIGHT

        if row_idx == len(members) + 2:
            for ci in range(1, 5 + days_in_month * 5):
                cell = ws.cell(row_idx, ci)
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=MED_SIDE,
                )

    ws.freeze_panes = 'E3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'PicaPonto_{date_from.strftime("%b%Y")}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

# ── POT routes ────────────────────────────────────────────────────────────

@rota_bp.route('/rota/hours/pot/draft', methods=['GET'])
@require_auth
def rota_hours_pot_draft():
    """Return live-computed hours for a team+month alongside any existing
    active POT record, consistency warnings, and a diff if POT exists.
    Management only."""
    if _get_rota_role(request.session) != 'management':
        return jsonify({'ok': False, 'error': 'Not authorised'}), 403

    team  = request.args.get('team', '').upper()
    month = request.args.get('month', '')   # YYYY-MM

    if team not in ('SOE', 'SOS'):
        return jsonify({'ok': False, 'error': 'team must be SOE or SOS'}), 400
    try:
        year, mo  = [int(x) for x in month.split('-')]
        date_from = date(year, mo, 1)
        date_to   = (date(year, mo + 1, 1) - timedelta(days=1)) if mo < 12 \
                    else date(year, 12, 31)
    except (ValueError, AttributeError):
        return jsonify({'ok': False, 'error': 'month must be YYYY-MM'}), 400

    leave_list = _load_json(LEAVE_FILE)
    if not isinstance(leave_list, list): leave_list = []
    published_overrides = _load_json(PUBLISHED_OVERRIDES_FILE)
    if not isinstance(published_overrides, list): published_overrides = []

    leave_map    = _build_leave_map(leave_list)
    override_map = _build_override_map(published_overrides)
    members      = _hr_teams().get(team, [])

    if not members:
        return jsonify({'ok': False, 'error': f'No members configured for {team}'}), 400
    computed = _compute_hours(date_from, date_to, members, leave_map, override_map)
    warnings = _check_hr_config_consistency()

    # Attach Employee IDs to computed results — sourced from person_directory.json,
    # since no POT commit exists yet at draft-preview time.
    emp_id_map = _LABEL_TO_RID
    computed_out = {}
    for name in members:
        h = computed.get(name, {'night_h': 0, 'ph_day_h': 0, 'ph_night_h': 0, 'ph_dates': []})
        computed_out[name] = {**h, 'employeeID': emp_id_map.get(name)}

    # Diff against existing active POT if one exists
    existing_pot = _active_pot_for(team, month)
    diff = None
    if existing_pot:
        diff = []
        pot_map = {e['name']: e for e in existing_pot['entries']}
        for name in members:
            comp = computed_out[name]
            pot_entry = pot_map.get(name)
            if not pot_entry:
                diff.append({'name': name, 'status': 'new_member'})
                continue
            fields_changed = {}
            for field in ('night_h', 'ph_day_h', 'ph_night_h'):
                comp_val = round(comp[field], 2)
                pot_val  = round(pot_entry[f'{field}_final'], 2)
                if abs(comp_val - pot_val) > 0.001:
                    fields_changed[field] = {'computed': comp_val, 'pot_final': pot_val}
            if fields_changed:
                diff.append({'name': name, 'status': 'changed', 'fields': fields_changed})
        # Members in POT but not in current hr_config
        for name in pot_map:
            if name not in members:
                diff.append({'name': name, 'status': 'removed_from_team'})

    return jsonify({
        'ok':          True,
        'team':        team,
        'month':       month,
        'computed':    computed_out,
        'member_order': members,
        'warnings':    warnings,
        'existing_pot': {
            'id':           existing_pot['id'],
            'committed_by': existing_pot['committed_by'],
            'committed_at': existing_pot['committed_at'],
            'supersedes':   existing_pot.get('supersedes'),
        } if existing_pot else None,
        'diff':        diff,
    })


@rota_bp.route('/rota/hours/pot/commit', methods=['POST'])
@require_auth
def rota_hours_pot_commit():
    """Commit a final set of hours values for a team+month to the POT.
    Accepts manually overridden final values alongside computed ones.
    Each overridden field requires a comment. Management only."""
    if _get_rota_role(request.session) != 'management':
        return jsonify({'ok': False, 'error': 'Not authorised'}), 403

    session    = request.session
    data       = request.get_json(silent=True) or {}
    team       = data.get('team', '').upper()
    month      = data.get('month', '')
    entries_in = data.get('entries', [])
    supersede_reason = data.get('supersede_reason', '').strip()

    if team not in ('SOE', 'SOS'):
        return jsonify({'ok': False, 'error': 'team must be SOE or SOS'}), 400
    try:
        year, mo  = [int(x) for x in month.split('-')]
        date_from = date(year, mo, 1)
        date_to   = (date(year, mo + 1, 1) - timedelta(days=1)) if mo < 12 \
                    else date(year, 12, 31)
    except (ValueError, AttributeError):
        return jsonify({'ok': False, 'error': 'month must be YYYY-MM'}), 400

    # Server-side recompute — never trust client-supplied "computed" values
    leave_list = _load_json(LEAVE_FILE)
    if not isinstance(leave_list, list): leave_list = []
    published_overrides = _load_json(PUBLISHED_OVERRIDES_FILE)
    if not isinstance(published_overrides, list): published_overrides = []

    leave_map    = _build_leave_map(leave_list)
    override_map = _build_override_map(published_overrides)
    members      = _hr_teams().get(team, [])
    display_names = _display_names()

    if not members:
        return jsonify({'ok': False, 'error': f'No members configured for {team}'}), 400

    computed = _compute_hours(date_from, date_to, members, leave_map, override_map)
    now      = _now_iso()

    # Check if a POT already exists — require supersede_reason if so
    records     = _load_pot()
    existing    = next((r for r in records
                        if r['team'] == team and r['month'] == month
                        and r['status'] == 'active'), None)
    if existing and not supersede_reason:
        return jsonify({
            'ok':    False,
            'error': 'A committed record already exists for this team and month. '
                     'Provide supersede_reason to replace it.',
            'existing_id': existing['id'],
            'needs_supersede_reason': True,
        }), 409

    # Validate and build final entries
    entries_by_name = {e.get('name'): e for e in entries_in}
    final_entries   = []
    warnings        = _check_hr_config_consistency()

    for name in members:
        comp     = computed.get(name, {'night_h': 0, 'ph_day_h': 0,
                                       'ph_night_h': 0, 'ph_dates': []})
        e_in     = entries_by_name.get(name, {})
        overrides_out = []

        entry = {
            'name':             name,
            'full_name':        display_names.get(name, name),
            'employeeID':       _LABEL_TO_RID.get(name),
            'ph_dates':         comp['ph_dates'],
            'night_h_computed':    round(comp['night_h'], 2),
            'ph_day_h_computed':   round(comp['ph_day_h'], 2),
            'ph_night_h_computed': round(comp['ph_night_h'], 2),
            'overrides':        [],
        }

        for field in ('night_h', 'ph_day_h', 'ph_night_h'):
            computed_val = round(comp[field], 2)
            # Client sends final value; fall back to computed if absent
            try:
                final_val = round(float(e_in.get(f'{field}_final', computed_val)), 2)
            except (TypeError, ValueError):
                final_val = computed_val

            entry[f'{field}_final'] = final_val

            if abs(final_val - computed_val) > 0.001:
                comment = (e_in.get('override_comments') or {}).get(field, '').strip()
                if not comment:
                    return jsonify({
                        'ok':    False,
                        'error': f'{name}: a comment is required for the '
                                 f'{field} override ({computed_val} → {final_val}).',
                    }), 400
                overrides_out.append({
                    'field':   field,
                    'from':    computed_val,
                    'to':      final_val,
                    'comment': comment,
                    'by':      session['username'],
                    'at':      now,
                })

        entry['overrides'] = overrides_out
        final_entries.append(entry)

    # Check that a superseding commit actually differs from the existing one
    if existing:
        existing_map = {e['name']: e for e in existing['entries']}
        any_diff = False
        for fe in final_entries:
            ex = existing_map.get(fe['name'])
            if not ex:
                any_diff = True
                break
            for field in ('night_h', 'ph_day_h', 'ph_night_h'):
                if abs(fe[f'{field}_final'] - ex.get(f'{field}_final', 0)) > 0.001:
                    any_diff = True
                    break
            if any_diff:
                break
        if not any_diff:
            return jsonify({
                'ok':    False,
                'error': 'The new values are identical to the existing committed '
                         'record. No superseding record was created.',
            }), 400

    # Supersede existing active record
    new_id = str(uuid.uuid4())[:8]
    if existing:
        existing['status']       = 'superseded'
        existing['superseded_by'] = new_id

    records.append({
        'id':              new_id,
        'team':            team,
        'month':           month,
        'status':          'active',
        'superseded_by':   None,
        'supersedes':      existing['id'] if existing else None,
        'supersede_reason': supersede_reason if existing else None,
        'computed_by':     session['username'],
        'committed_by':    session['username'],
        'committed_at':    now,
        'consistency_warnings': warnings,
        'entries':         final_entries,
    })
    _save_pot(records)

    return jsonify({
        'ok':          True,
        'id':          new_id,
        'superseded':  existing['id'] if existing else None,
        'warnings':    warnings,
    })


@rota_bp.route('/rota/hours/pot', methods=['GET'])
@require_auth
def rota_hours_pot_get():
    """Return POT records for a team+month. Active record first, then
    superseded chain in reverse chronological order.
    Accessible to all authenticated users (read-only)."""
    team  = request.args.get('team', '').upper()
    month = request.args.get('month', '')
    if not team or not month:
        return jsonify({'ok': False, 'error': 'team and month required'}), 400
    if team not in ('SOE', 'SOS'):
        return jsonify({'ok': False, 'error': 'team must be SOE or SOS'}), 400

    all_records = [r for r in _load_pot()
                   if r['team'] == team and r['month'] == month]
    # Active first, then superseded newest-first
    active     = [r for r in all_records if r['status'] == 'active']
    superseded = sorted([r for r in all_records if r['status'] == 'superseded'],
                        key=lambda r: r['committed_at'], reverse=True)

    return jsonify({
        'ok':      True,
        'team':    team,
        'month':   month,
        'records': active + superseded,
    })


@rota_bp.route('/rota/hours/debug', methods=['GET'])
@require_auth
def rota_hours_debug():
    if _get_rota_role(request.session) != 'management':
        return jsonify({'ok': False, 'error': 'Not authorised'}), 403

    hr_teams    = _hr_teams()
    sos_members = hr_teams.get('SOS', [])
    soe_members = hr_teams.get('SOE', [])

    known = list(set(MANAGEMENT_SHIFTS) | set(ENGINEERING_OFFSETS) | set(SPECIALIST_OFFSETS))

    leave_list = _load_json(LEAVE_FILE)
    if not isinstance(leave_list, list): leave_list = []
    published_overrides = _load_json(PUBLISHED_OVERRIDES_FILE)
    if not isinstance(published_overrides, list): published_overrides = []

    leave_map    = _build_leave_map(leave_list)
    override_map = _build_override_map(published_overrides)

    test_month_from = date(2026, 6, 1)
    test_month_to   = date(2026, 6, 30)
    all_members = sos_members + soe_members
    hours = _compute_hours(test_month_from, test_month_to,
                           all_members, leave_map, override_map)

    # Spot-check the first SOS specialist on 3 days (no name hardcoded here —
    # picked dynamically so this endpoint never needs a real name in source)
    spot_name = next((n for n in sos_members if n in SPECIALIST_OFFSETS), None)
    spot = {}
    for d_str in ['2026-06-01', '2026-06-02', '2026-06-04']:
        d = date.fromisoformat(d_str)
        if spot_name:
            spot[f'{spot_name}@{d_str}'] = {
                'effective': _effective_shift_for_hours(spot_name, d, leave_map, override_map),
                'resolved':  _resolve_shift(spot_name, d, leave_map, override_map),
                'base':      _base_shift(spot_name, d),
            }

    return jsonify({
        'sos_members':      sos_members,
        'soe_members':      soe_members,
        'known_rota_names': sorted(known),
        'hours_keys':       list(hours.keys()),
        'hours_first_sos':  hours.get(sos_members[0]) if sos_members else None,
        'spot_checks':      spot,
    })

# ════════════════════════════════════════════════════════════════════════════
#  AL ALLOWANCE
# ════════════════════════════════════════════════════════════════════════════

@rota_bp.route('/rota/al-allowance', methods=['GET'])
@require_auth
def rota_al_allowance_get():
    session   = request.session
    rota_role = _get_rota_role(session)
    try:
        year = int(request.args.get('year', date.today().year))
    except ValueError:
        return jsonify({'ok': False, 'error': 'year must be an integer'}), 400

    al = _load_al_file()
    leave_list = _load_json(LEAVE_FILE)
    if not isinstance(leave_list, list):
        leave_list = []
    leave_map = _build_leave_map(leave_list)

    all_names = (list(MANAGEMENT_SHIFTS) +
                 list(ENGINEERING_OFFSETS) +
                 list(SPECIALIST_OFFSETS))

    if rota_role == 'management':
        names = all_names
    elif rota_role == 'staff':
        my_name = _rota_display_name(session['username'])
        names = [my_name] if my_name in all_names else []
    else:
        return jsonify({'ok': False, 'error': 'Not authorised'}), 403

    balances = {n: _compute_al_balance(al, n, year, leave_list, leave_map) for n in names}
    _save_al_file(al)  # persist any evaluation defaults just computed

    return jsonify({
        'ok': True, 'year': year,
        'mhd_default_days': al['yearly'].get(str(year), {}).get('mhd_default_days'),
        'balances': balances,
    })


@rota_bp.route('/rota/al-allowance', methods=['PUT'])
@require_auth
def rota_al_allowance_put():
    err = _require_management()
    if err: return err

    session = request.session
    data    = request.get_json(silent=True) or {}
    name    = data.get('name', '').strip()
    year    = data.get('year')
    field   = data.get('field', '').strip()   # 'base_allowance' | 'mhd' | 'absence'
    final_days = data.get('final_days')
    reason  = (data.get('override_reason') or '').strip()

    if field not in ('base_allowance', 'mhd', 'absence'):
        return jsonify({'ok': False, 'error': 'Invalid field'}), 400
    if field == 'absence':
        return jsonify({'ok': False,
                        'error': 'absence is auto-computed and not directly editable via this field'}), 400
    if not name or year is None:
        return jsonify({'ok': False, 'error': 'name and year required'}), 400
    try:
        year = int(year)
        final_days = float(final_days)
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'year and final_days must be numeric'}), 400
    if not reason:
        return jsonify({'ok': False, 'error': 'override_reason is required for manual entries'}), 400

    al = _load_al_file()
    if name not in al['members']:
        return jsonify({'ok': False, 'error': f'{name} has no join_date on file — set that first'}), 400

    entry = _evaluate_member_year(al, name, year)
    entry[field]['auto_or_manual']  = 'manual'
    entry[field]['final_days']      = final_days
    entry[field]['override_reason'] = reason

    al['yearly'][str(year)]['members'][name] = entry
    _save_al_file(al)
    return jsonify({'ok': True, 'entry': entry})


@rota_bp.route('/rota/al-allowance/join-date', methods=['PUT'])
@require_auth
def rota_al_join_date_put():
    err = _require_management()
    if err: return err
    data = request.get_json(silent=True) or {}
    name = data.get('name', '').strip()
    jd   = data.get('join_date', '').strip()
    if not name or not jd:
        return jsonify({'ok': False, 'error': 'name and join_date required'}), 400
    try:
        date.fromisoformat(jd)
    except ValueError:
        return jsonify({'ok': False, 'error': 'Invalid date format'}), 400

    al = _load_al_file()
    al['members'].setdefault(name, {})['join_date'] = jd
    _save_al_file(al)
    return jsonify({'ok': True})


@rota_bp.route('/rota/al-allowance/config', methods=['PUT'])
@require_auth
def rota_al_config_put():
    """Set the company-wide MHD default for a year and re-evaluate MHD
    across all eligible (join_date <= N-1) members."""
    err = _require_management()
    if err: return err
    data = request.get_json(silent=True) or {}
    year = data.get('year')
    mhd  = data.get('mhd_default_days')
    try:
        year = int(year)
        mhd  = float(mhd)
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'year and mhd_default_days must be numeric'}), 400

    al = _load_al_file()
    yr_block = al['yearly'].setdefault(str(year), {})
    yr_block['mhd_default_days'] = mhd

    all_names = (list(MANAGEMENT_SHIFTS) +
                 list(ENGINEERING_OFFSETS) +
                 list(SPECIALIST_OFFSETS))
    for n in all_names:
        if n in al['members']:
            _evaluate_member_year(al, n, year)

    _save_al_file(al)
    return jsonify({'ok': True, 'mhd_default_days': mhd})


@rota_bp.route('/rota/al-allowance/misc', methods=['POST'])
@require_auth
def rota_al_misc_post():
    err = _require_management()
    if err: return err
    session = request.session
    data    = request.get_json(silent=True) or {}
    name    = data.get('name', '').strip()
    year    = data.get('year')
    hours   = data.get('hours')
    reason  = (data.get('reason') or '').strip()
    date_s  = data.get('date', '').strip()

    if not name or year is None or not reason or not date_s:
        return jsonify({'ok': False, 'error': 'name, year, hours, date, reason required'}), 400
    try:
        year  = int(year)
        hours = float(hours)
        date.fromisoformat(date_s)
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'Invalid year, hours, or date'}), 400

    al    = _load_al_file()
    entry = _evaluate_member_year(al, name, year)
    entry.setdefault('misc_entries', []).append({
        'id':         str(uuid.uuid4())[:8],
        'hours':      hours,
        'reason':     reason,
        'date':       date_s,
        'created_by': session['username'],
        'created_at': _now_iso(),
    })
    al['yearly'][str(year)]['members'][name] = entry
    _save_al_file(al)
    return jsonify({'ok': True, 'entry': entry})


@rota_bp.route('/rota/al-allowance/misc/<entry_id>', methods=['DELETE'])
@require_auth
def rota_al_misc_delete(entry_id):
    err = _require_management()
    if err: return err
    data = request.get_json(silent=True) or {}
    name = data.get('name', '').strip()
    year = data.get('year')
    if not name or year is None:
        return jsonify({'ok': False, 'error': 'name and year required'}), 400
    try:
        year = int(year)
    except ValueError:
        return jsonify({'ok': False, 'error': 'Invalid year'}), 400

    al    = _load_al_file()
    entry = al.get('yearly', {}).get(str(year), {}).get('members', {}).get(name)
    if not entry:
        return jsonify({'ok': False, 'error': 'No entry found'}), 404
    before = len(entry.get('misc_entries', []))
    entry['misc_entries'] = [e for e in entry.get('misc_entries', []) if e['id'] != entry_id]
    if len(entry['misc_entries']) == before:
        return jsonify({'ok': False, 'error': 'Misc entry not found'}), 404

    _save_al_file(al)
    return jsonify({'ok': True, 'entry': entry})

# ════════════════════════════════════════════════════════════════════════════
#  SOE WEEKEND COVERAGE
# ════════════════════════════════════════════════════════════════════════════

def _is_working_soe_shift(shift: str) -> bool:
    """Return True if the shift counts as worked for SOE weekend coverage.
    Any resolved shift that isn't OFF/leave/absent qualifies."""
    if not shift or shift == 'OFF':
        return False
    if shift in ('PARENTAL', 'MARITAL'):
        return False
    if shift.startswith('AL_') or shift.startswith('ABSENT'):
        return False
    # AL_APPROVED|0900-1800 style
    if '|' in shift:
        prefix = shift.split('|', 1)[0]
        if prefix in ('AL_APPROVED', 'AL_PENDING', 'ABSENT'):
            return False
        # e.g. ABSENT|0900-1800 — not working
        return True
    return True


def _soe_weekend_counts(year: int,
                        leave_map: dict,
                        override_map: dict,
                        soe_join_dates: dict) -> dict:
    """
    For each SOE member, count Saturday and Sunday days worked in `year`,
    respecting join date. Also returns available weekend days (Sat+Sun from
    join date to end of year) for ratio computation.

    Returns dict: name -> {
        'worked': int,
        'available': int,
        'ratio': float,          # worked / available, 0.0 if no available days
        'join_date': str | None,
    }
    """
    year_start = date(year, 1, 1)
    year_end   = date(year, 12, 31)

    results = {}
    for name in ENGINEERING_OFFSETS:
        jd_str = soe_join_dates.get(name)
        try:
            join = date.fromisoformat(jd_str) if jd_str else year_start
        except ValueError:
            join = year_start

        # Clamp: don't count days before join date, don't go past year end
        effective_start = max(join, year_start)

        worked    = 0
        available = 0

        d = effective_start
        while d <= year_end:
            if d.weekday() in (5, 6):  # Saturday=5, Sunday=6
                available += 1
                shift = _resolve_shift(name, d, leave_map, override_map)
                if _is_working_soe_shift(shift):
                    worked += 1
            d += timedelta(days=1)

        results[name] = {
            'worked':    worked,
            'available': available,
            'ratio':     round(worked / available, 4) if available else 0.0,
            'join_date': jd_str,
        }

    return results


def _soe_aggregate(years: list[int],
                   leave_map: dict,
                   override_map: dict,
                   soe_join_dates: dict) -> dict:
    """
    Sum worked and available across all requested years per person.
    Returns same shape as _soe_weekend_counts but totalled.
    """
    totals = {name: {'worked': 0, 'available': 0} for name in ENGINEERING_OFFSETS}

    for year in years:
        yearly = _soe_weekend_counts(year, leave_map, override_map, soe_join_dates)
        for name, data in yearly.items():
            totals[name]['worked']    += data['worked']
            totals[name]['available'] += data['available']

    results = {}
    for name, t in totals.items():
        results[name] = {
            'worked':    t['worked'],
            'available': t['available'],
            'ratio':     round(t['worked'] / t['available'], 4) if t['available'] else 0.0,
            'join_date': soe_join_dates.get(name),
        }
    return results


def _soe_team_delta(counts: dict) -> dict:
    """Add delta_from_mean (worked days vs team mean) to each entry."""
    worked_values = [v['worked'] for v in counts.values()]
    if not worked_values:
        return counts
    mean = sum(worked_values) / len(worked_values)
    for name in counts:
        counts[name]['delta'] = round(counts[name]['worked'] - mean, 2)
    return counts


@rota_bp.route('/rota/soe-weekends', methods=['GET'])
@require_auth
def rota_soe_weekends():
    """
    SOE weekend coverage counts.
    Query params:
      year=YYYY        — single year (defaults to current year)
      aggregate=1      — return aggregate across all years from earliest join to now
    Accessible to management and engineering roles only.
    """
    session   = request.session
    rota_role = _get_rota_role(session)
    user_role = session.get('role', '')

    if rota_role != 'management' and user_role != 'engineer':
        return jsonify({'ok': False, 'error': 'Not authorised'}), 403

    today        = date.today()
    current_year = today.year

    # Load shared state once
    leave_list = _load_json(LEAVE_FILE)
    if not isinstance(leave_list, list):
        leave_list = []
    published_overrides = _load_json(PUBLISHED_OVERRIDES_FILE)
    if not isinstance(published_overrides, list):
        published_overrides = []

    leave_map    = _build_leave_map(leave_list)
    override_map = _build_override_map(published_overrides)

    soe_join_dates = {v['rota_label']: v['soe_join_date']
                       for v in _DIR.values() if v.get('soe_join_date')}

    do_aggregate = request.args.get('aggregate', '0') == '1'

    if do_aggregate:
        # Determine earliest join year across all SOE members
        join_years = []
        for name in ENGINEERING_OFFSETS:
            jd_str = soe_join_dates.get(name)
            if jd_str:
                try:
                    join_years.append(date.fromisoformat(jd_str).year)
                except ValueError:
                    pass
        earliest = min(join_years) if join_years else current_year
        years    = list(range(earliest, current_year + 1))

        aggregate = _soe_aggregate(years, leave_map, override_map, soe_join_dates)
        aggregate = _soe_team_delta(aggregate)

        # Also return per-year breakdown for the dropdown
        yearly_breakdown = {}
        for y in years:
            yc = _soe_weekend_counts(y, leave_map, override_map, soe_join_dates)
            yearly_breakdown[str(y)] = _soe_team_delta(yc)

        return jsonify({
            'ok':               True,
            'mode':             'aggregate',
            'years':            years,
            'aggregate':        aggregate,
            'yearly_breakdown': yearly_breakdown,
            'members':          list(ENGINEERING_OFFSETS.keys()),
        })

    else:
        try:
            year = int(request.args.get('year', current_year))
        except ValueError:
            return jsonify({'ok': False, 'error': 'year must be an integer'}), 400

        counts = _soe_weekend_counts(year, leave_map, override_map, soe_join_dates)
        counts = _soe_team_delta(counts)

        return jsonify({
            'ok':     True,
            'mode':   'year',
            'year':   year,
            'counts': counts,
            'members': list(ENGINEERING_OFFSETS.keys()),
        })

@rota_bp.route('/rota/feedback', methods=['POST'])
@require_auth
def rota_feedback_post():
    session = request.session
    data    = request.get_json(silent=True) or {}
    fb_type = data.get('type', '').strip()
    title   = data.get('title', '').strip()
    body    = data.get('body', '').strip()

    if fb_type not in ('bug', 'enhancement'):
        return jsonify({'ok': False, 'error': 'Invalid feedback type'}), 400
    if not title:
        return jsonify({'ok': False, 'error': 'Title required'}), 400
    if not body:
        return jsonify({'ok': False, 'error': 'Description required'}), 400

    feedback = _load_json(FEEDBACK_FILE)
    if not isinstance(feedback, dict):
        feedback = {}
    feedback.setdefault('bug', [])
    feedback.setdefault('enhancement', [])

    feedback[fb_type].append({
        'id':           str(uuid.uuid4())[:8],
        'submitted_by': session['username'],
        'submitted_at': _now_iso(),
        'title':        title,
        'body':         body,
        'status':       'unreviewed',
        'actioned_by':  None,
        'actioned_at':  None,
    })

    _save_json(FEEDBACK_FILE, feedback)
    return jsonify({'ok': True})


@rota_bp.route('/rota/feedback', methods=['GET'])
@require_auth
def rota_feedback_get():
    if _get_rota_role(request.session) != 'management':
        return jsonify({'ok': False, 'error': 'Not authorised'}), 403
    feedback = _load_json(FEEDBACK_FILE)
    if not isinstance(feedback, dict):
        feedback = {'bug': [], 'enhancement': []}
    feedback.setdefault('bug', [])
    feedback.setdefault('enhancement', [])
    return jsonify({'ok': True, 'feedback': feedback})


@rota_bp.route('/rota/feedback/<feedback_id>', methods=['PUT'])
@require_auth
def rota_feedback_put(feedback_id):
    if _get_rota_role(request.session) != 'management':
        return jsonify({'ok': False, 'error': 'Not authorised'}), 403

    session    = request.session
    data       = request.get_json(silent=True) or {}
    new_status = data.get('status', '').strip()

    if new_status not in ('resolved', 'dismissed'):
        return jsonify({'ok': False, 'error': 'status must be resolved or dismissed'}), 400

    feedback = _load_json(FEEDBACK_FILE)
    if not isinstance(feedback, dict):
        return jsonify({'ok': False, 'error': 'No feedback data'}), 500

    for fb_type in ('bug', 'enhancement'):
        for entry in feedback.get(fb_type, []):
            if entry.get('id') == feedback_id:
                entry['status']      = new_status
                entry['actioned_by'] = session['username']
                entry['actioned_at'] = _now_iso()
                _save_json(FEEDBACK_FILE, feedback)
                return jsonify({'ok': True})

    return jsonify({'ok': False, 'error': 'Feedback entry not found'}), 404

def register_routes(app) -> None:
    app.register_blueprint(rota_bp)